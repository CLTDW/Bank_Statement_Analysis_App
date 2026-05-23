from typing import Optional, List, Dict
from pathlib import Path
import pandas as pd
import logging
import os
import re
import time

from src.utils import (
    standardize_header,
    optimize_duplicate_columns,
    detect_file_encoding,
    detect_file_delimiter,
    preprocess_txt_file,
    is_excel_encrypted
)
from src.utils.multi_account_handler import MultiAccountHandler


class FileProcessor:
    """
    文件处理器类，负责处理不同类型的银行流水文件
    """
    def __init__(self, config: Dict, mapping_manager, object_names: List[str] = None):
        """
        初始化文件处理器
        
        Args:
            config: 配置字典
            mapping_manager: 映射管理器实例
            object_names: 用户输入的对象名称列表
        """
        self.config = config
        self.mapping_manager = mapping_manager
        self.logger = logging.getLogger('bank_statement_aggregator')
        
        # 获取映射关系
        self.header_mapping = mapping_manager.header_mapping
        self.keyword_mapping = mapping_manager.keyword_mapping
        self.unified_headers = mapping_manager.expected_columns
        # 获取排除规则
        self.exclusion_rules = getattr(mapping_manager, 'get_exclusion_rules', lambda: {} )()
        
        # 初始化多账户处理器
        # 获取姓氏库
        surnames = getattr(mapping_manager, 'get_surnames', lambda: {'single': [], 'compound': []})()
        self.multi_account_handler = MultiAccountHandler(self.header_mapping, self.keyword_mapping, object_names, surnames)
        
        # 缓存：避免重复读取同一个Excel文件
        self.excel_file_cache: Dict[str, pd.ExcelFile] = {}
        # 缓存：避免重复检测编码
        self.encoding_cache: Dict[str, str] = {}
        # 缓存：避免重复检测分隔符
        self.delimiter_cache: Dict[str, str] = {}
        # 缓存：避免重复计算表头得分
        self.header_score_cache: Dict[tuple, float] = {}
        # 缓存：避免重复处理相同的文件
        self.file_result_cache: Dict[str, Optional[pd.DataFrame]] = {}
        # 缓存大小限制
        self.cache_size_limit = config.get('cache_size_limit', 100)
        # 缓存访问时间，用于LRU缓存清理
        self.cache_access_times: Dict[str, float] = {}
        # 缓存创建时间，用于过期清理
        self.cache_creation_times: Dict[str, float] = {}
        # 缓存过期时间（秒）
        self.cache_expiration_times = {
            'excel_file': 300,    # 5分钟
            'encoding': 3600,      # 1小时
            'delimiter': 3600,     # 1小时
            'header': 1800,        # 30分钟
            'file': 600            # 10分钟
        }
        # 内存使用监控
        self.memory_threshold = 80  # 内存使用阈值（百分比）
        self.last_memory_check = time.time()
        self.memory_check_interval = 10  # 内存检查间隔（秒）
        # 大文件阈值（默认10MB）
        self.large_file_threshold = config.get('large_file_threshold', 10 * 1024 * 1024)
        # 分块大小（默认10000行）
        self.chunk_size = config.get('chunk_size', 10000)
    
    @staticmethod
    def _excel_col_to_num(col: str) -> int:
        """
        将Excel列字母转换为数字索引（A=1, B=2, ..., Z=26, AA=27, 等等）
        
        Args:
            col: Excel列字母，如 "A", "B", "AA"
            
        Returns:
            对应的数字索引（1-based）
        """
        num = 0
        for c in col.upper():
            num = num * 26 + (ord(c) - ord('A') + 1)
        return num
    
    def _update_cache_access(self, cache_type: str, key: str) -> None:
        """
        更新缓存访问时间
        
        Args:
            cache_type: 缓存类型
            key: 缓存键
        """
        import time
        current_time = time.time()
        # 使用统一的键格式
        cache_key = f"{cache_type}:{key}"
        self.cache_access_times[cache_key] = current_time
        
        # 记录缓存创建时间（如果是新缓存）
        if cache_key not in self.cache_creation_times:
            self.cache_creation_times[cache_key] = current_time
        
        # 定期检查内存使用情况
        if current_time - self.last_memory_check > self.memory_check_interval:
            self._check_memory_usage()
            self.last_memory_check = current_time
    
    @staticmethod
    def _get_file_cache_key(file_path: str) -> str:
        """
        获取文件缓存键，包含文件路径和修改时间
        
        Args:
            file_path: 文件路径
            
        Returns:
            缓存键
        """
        try:
            mtime = os.path.getmtime(file_path)
            return f"{file_path}:{mtime}"
        except Exception:
            return file_path
    
    def _adjust_cache_size_dynamically(self):
        """
        根据系统内存情况动态调整缓存大小
        """
        try:
            import psutil
            memory = psutil.virtual_memory()
            memory_usage = memory.percent
            
            # 根据内存使用情况调整缓存大小
            if memory_usage > 80:
                # 内存紧张，大幅减少缓存
                self.cache_size_limit = max(20, int(self.config.get('cache_size_limit', 100) * 0.2))
                self.logger.warning(f"【缓存管理】内存紧张，调整缓存大小为：{self.cache_size_limit}")
            elif memory_usage > 60:
                # 内存适中，适度减少缓存
                self.cache_size_limit = max(50, int(self.config.get('cache_size_limit', 100) * 0.5))
            else:
                # 内存充足，使用默认缓存大小
                self.cache_size_limit = self.config.get('cache_size_limit', 100)
        except:
            # 如果无法获取内存信息，使用默认值
            self.cache_size_limit = self.config.get('cache_size_limit', 100)
    
    def _cleanup_cache(self) -> None:
        """
        清理缓存，当缓存大小超过限制时
        """
        import time
        current_time = time.time()
        
        # 动态调整缓存大小
        self._adjust_cache_size_dynamically()
        
        # 检查并清理表头得分缓存
        if len(self.header_score_cache) > self.cache_size_limit:
            # 清理过期缓存
            valid_items = []
            for key, value in self.header_score_cache.items():
                cache_key = f"header:{str(key)}"
                creation_time = self.cache_creation_times.get(cache_key, 0)
                if current_time - creation_time < self.cache_expiration_times.get('header', 1800):
                    valid_items.append((key, value))
            
            # 按访问时间排序，保留最近使用的缓存
            sorted_items = sorted(valid_items, 
                                key=lambda x: self.cache_access_times.get(f"header:{str(x[0])}", 0), 
                                reverse=True)
            # 保留前80%的缓存
            keep_count = int(self.cache_size_limit * 0.8)
            self.header_score_cache = dict(sorted_items[:keep_count])
            # 清理访问时间记录
            self._cleanup_cache_access_times("header", [str(key) for key in self.header_score_cache.keys()])

        
        # 检查并清理文件结果缓存
        if len(self.file_result_cache) > self.cache_size_limit:
            # 清理过期缓存
            valid_items = []
            for key, value in self.file_result_cache.items():
                cache_key = f"file:{key}"
                creation_time = self.cache_creation_times.get(cache_key, 0)
                if current_time - creation_time < self.cache_expiration_times.get('file', 600):
                    valid_items.append((key, value))
            
            # 按访问时间排序，保留最近使用的缓存
            sorted_items = sorted(valid_items, 
                                key=lambda x: self.cache_access_times.get(f"file:{x[0]}", 0), 
                                reverse=True)
            # 保留前80%的缓存
            keep_count = int(self.cache_size_limit * 0.8)
            self.file_result_cache = dict(sorted_items[:keep_count])
            # 清理访问时间记录
            self._cleanup_cache_access_times("file", list(self.file_result_cache.keys()))

        
        # 检查并清理Excel文件缓存
        if len(self.excel_file_cache) > self.cache_size_limit:
            # 清理过期缓存
            valid_items = []
            for key, value in self.excel_file_cache.items():
                cache_key = f"excel_file:{key}"
                creation_time = self.cache_creation_times.get(cache_key, 0)
                if current_time - creation_time < self.cache_expiration_times.get('excel_file', 300):
                    valid_items.append((key, value))
            
            # 按访问时间排序，保留最近使用的缓存
            sorted_items = sorted(valid_items, 
                                key=lambda x: self.cache_access_times.get(f"excel_file:{x[0]}", 0), 
                                reverse=True)
            # 保留前80%的缓存
            keep_count = int(self.cache_size_limit * 0.8)
            self.excel_file_cache = dict(sorted_items[:keep_count])
            # 清理访问时间记录
            self._cleanup_cache_access_times("excel_file", list(self.excel_file_cache.keys()))

        
        # 检查并清理编码和分隔符缓存
        if len(self.encoding_cache) > self.cache_size_limit:
            # 清理过期缓存
            valid_items = []
            for key, value in self.encoding_cache.items():
                cache_key = f"encoding:{key}"
                creation_time = self.cache_creation_times.get(cache_key, 0)
                if current_time - creation_time < self.cache_expiration_times.get('encoding', 3600):
                    valid_items.append((key, value))
            
            # 按访问时间排序，保留最近使用的缓存
            sorted_items = sorted(valid_items, 
                                key=lambda x: self.cache_access_times.get(f"encoding:{x[0]}", 0), 
                                reverse=True)
            # 保留前80%的缓存
            keep_count = int(self.cache_size_limit * 0.8)
            self.encoding_cache = dict(sorted_items[:keep_count])
            # 清理访问时间记录
            self._cleanup_cache_access_times("encoding", list(self.encoding_cache.keys()))

        
        if len(self.delimiter_cache) > self.cache_size_limit:
            # 清理过期缓存
            valid_items = []
            for key, value in self.delimiter_cache.items():
                cache_key = f"delimiter:{key}"
                creation_time = self.cache_creation_times.get(cache_key, 0)
                if current_time - creation_time < self.cache_expiration_times.get('delimiter', 3600):
                    valid_items.append((key, value))
            
            # 按访问时间排序，保留最近使用的缓存
            sorted_items = sorted(valid_items, 
                                key=lambda x: self.cache_access_times.get(f"delimiter:{x[0]}", 0), 
                                reverse=True)
            # 保留前80%的缓存
            keep_count = int(self.cache_size_limit * 0.8)
            self.delimiter_cache = dict(sorted_items[:keep_count])
            # 清理访问时间记录
            self._cleanup_cache_access_times("delimiter", list(self.delimiter_cache.keys()))

    
    def _cleanup_cache_access_times(self, cache_type: str, valid_keys: list) -> None:
        """
        清理缓存访问时间记录
        
        Args:
            cache_type: 缓存类型
            valid_keys: 有效的缓存键列表
        """
        keys_to_remove = []
        for key in self.cache_access_times:
            if key.startswith(f"{cache_type}:"):
                # 提取原始键
                original_key = key.split(":", 1)[1]
                if original_key not in valid_keys:
                    keys_to_remove.append(key)
        # 删除无效的访问时间记录
        for key in keys_to_remove:
            del self.cache_access_times[key]
            # 同时清理创建时间记录
            if key in self.cache_creation_times:
                del self.cache_creation_times[key]
    
    def _check_memory_usage(self):
        """
        检查内存使用情况，当超过阈值时进行内存清理
        """
        try:
            import psutil
            memory = psutil.virtual_memory()
            memory_usage = memory.percent
            

            
            if memory_usage > self.memory_threshold:
                self.logger.warning(f"【内存监控】内存使用超过阈值，进行紧急内存清理")
                
                # 紧急清理所有缓存
                self.excel_file_cache.clear()
                self.encoding_cache.clear()
                self.delimiter_cache.clear()
                self.header_score_cache.clear()
                self.file_result_cache.clear()
                self.cache_access_times.clear()
                self.cache_creation_times.clear()
                
                # 强制垃圾回收
                import gc
                gc.collect()
                
                # 再次检查内存使用情况
                memory = psutil.virtual_memory()
                new_memory_usage = memory.percent
                self.logger.info(f"【内存监控】紧急清理后内存使用：{new_memory_usage:.2f}%")
        except:
            pass  # 如果无法获取内存信息，忽略
    
    def _calculate_header_score(self, standardized_headers: List[str]) -> float:
        """
        计算表头得分
        
        Args:
            standardized_headers: 标准化后的表头列表
            
        Returns:
            表头得分
        """
        # 使用缓存避免重复计算
        headers_tuple = tuple(standardized_headers)
        if headers_tuple in self.header_score_cache:
            # 更新访问时间
            self._update_cache_access("header", str(headers_tuple))
            return self.header_score_cache[headers_tuple]
        
        exact_match_score = sum(1 for h in standardized_headers if h in self.header_mapping)
        keyword_match_score = 0
        for h in standardized_headers:
            for unified_col, keywords in self.keyword_mapping.items():
                if any(keyword in h for keyword in keywords):
                    keyword_match_score += 1
                    break
        
        score = exact_match_score * 2 + keyword_match_score * 1
        # 缓存结果
        self.header_score_cache[headers_tuple] = score
        # 更新访问时间
        self._update_cache_access("header", str(headers_tuple))
        # 检查缓存大小
        self._cleanup_cache()
        return score
    
    def _is_large_file(self, file_path: str) -> bool:
        """
        判断文件是否为大文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            是否为大文件
        """
        try:
            file_size = os.path.getsize(file_path)
            return file_size > self.large_file_threshold
        except Exception:
            return False
    
    def _process_large_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        分块处理大文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件处理】开始分块处理：{file_name}")
        
        try:
            # 导入文件类型检测函数
            from src.utils.file_utils import detect_file_type, get_file_extension_from_mime, magic_available, is_excel_file, is_text_file, is_pdf_file
            
            # 检测文件类型
            file_type = detect_file_type(file_path)
            file_ext = get_file_extension_from_mime(file_type)
            
            # 首先检查是否需要多账户处理或无户名单账户处理
            need_multi_account, best_header_score, need_single_account_no_name = self._need_multi_account_processing(file_path)
            if need_multi_account:
                self.logger.info(f"【大文件处理】检测到需要多账户处理，切换到多账户处理流程")
                return self._process_large_multi_account_file(file_path, best_header_score)
            elif need_single_account_no_name:
                self.logger.info(f"【大文件处理】检测到需要无户名单账户处理，切换到无户名单账户处理流程")
                return self._process_large_single_account_no_name(file_path)
            
            # 根据文件类型选择处理方式
            if is_excel_file(file_path):
                # 处理大型Excel文件
                return self._process_large_excel_file(file_path)
            elif is_text_file(file_path):
                # 处理大型文本文件（CSV/TXT/HTML）
                return self._process_large_text_file(file_path)
            else:
                # 其他文件类型，使用常规处理
                self.logger.info(f"【大文件处理】{file_name} 类型未知，使用常规处理")
                return self._process_text_file(file_path)
                
        except Exception as e:
            self.logger.error(f"【大文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _check_header_for_account_name(self, header_df: pd.DataFrame) -> tuple:
        """
        检查表头是否缺少户主名称
        
        Args:
            header_df: 表头数据
            
        Returns:
            (是否缺少户主名称, 最佳表头行评分, 是否存在其他表头行评分大于等于最佳表头行评分)
        """
        # 寻找最佳表头
        current_best_header_row = 0
        current_best_header_score = 0
        
        for i in range(min(50, len(header_df))):
            header_candidate = header_df.iloc[i]
            standardized_headers = [standardize_header(str(h)) for h in header_candidate]
            score = self._calculate_header_score(standardized_headers)
            
            if score > current_best_header_score:
                current_best_header_score = score
                current_best_header_row = i
        
        # 检查是否存在其他表头行评分大于等于最佳表头行评分
        has_other_header_rows = False
        for i in range(min(50, len(header_df))):
            if i == current_best_header_row:
                continue
            header_candidate = header_df.iloc[i]
            standardized_headers = [standardize_header(str(h)) for h in header_candidate]
            score = self._calculate_header_score(standardized_headers)
            if score >= current_best_header_score:
                has_other_header_rows = True
                break
        
        # 检查最佳表头是否缺少户主名称
        has_missing_account_name = True
        if current_best_header_row < len(header_df):
            best_header = header_df.iloc[current_best_header_row].apply(standardize_header)
            # 创建临时映射函数，检查是否缺少户主名称
            def temp_map(df_temp):
                # 表头标准化
                df_temp.columns = [standardize_header(col) for col in df_temp.columns]
                standardized_headers = list(df_temp.columns)
                
                # 映射到统一表头
                mapped_df = pd.DataFrame(columns=self.unified_headers)
                matched_cols = []
                
                # 收集所有可能的匹配
                all_matches = []
                
                # 1. 收集所有精确匹配
                for col in standardized_headers:
                    if col in self.header_mapping:
                        unified_col = self.header_mapping[col]
                        mapping_items = list(self.header_mapping.items())
                        position = next((i for i, (k, v) in enumerate(mapping_items) if k == col), len(mapping_items))
                        position_weight = max(0, 10 - position * 0.1)
                        weight = 20 + position_weight
                        
                        all_matches.append({
                            'original': col,
                            'matched': unified_col,
                            'method': '精确匹配',
                            'weight': weight,
                            'position': position
                        })
                
                # 2. 收集所有关键词匹配
                for col in standardized_headers:
                    if any(m['original'] == col and m['method'] == '精确匹配' for m in all_matches):
                        continue
                    
                    for unified_col, keywords in self.keyword_mapping.items():
                        for i, keyword in enumerate(keywords):
                            if unified_col in self.exclusion_rules:
                                exclusion_keywords = self.exclusion_rules[unified_col]
                                if any(exclusion_keyword in col for exclusion_keyword in exclusion_keywords):
                                    continue
                            
                            score = 0
                            if col == keyword:
                                score = 10
                            elif col.startswith(keyword) or col.endswith(keyword):
                                score = 8
                            elif keyword in col:
                                keyword_chars = set(keyword)
                                col_chars = set(col)
                                if keyword_chars:
                                    common_chars = keyword_chars.intersection(col_chars)
                                    coverage_ratio = len(common_chars) / len(keyword_chars)
                                    if coverage_ratio > 0.5:
                                        score = 5
                            
                            if score > 0:
                                keyword_position_weight = max(0, 5 - i * 0.5)
                                weight = score + keyword_position_weight
                                
                                all_matches.append({
                                    'original': col,
                                    'matched': unified_col,
                                    'method': '关键词匹配',
                                    'keyword': keyword,
                                    'weight': weight,
                                    'score': score,
                                    'keyword_position': i
                                })
                
                # 处理匹配冲突
                matches_by_original = {}
                for match in all_matches:
                    original_col = match['original']
                    if original_col not in matches_by_original:
                        matches_by_original[original_col] = []
                    matches_by_original[original_col].append(match)
                
                selected_matches = []
                for original_col, matches in matches_by_original.items():
                    sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                    best_match = sorted_matches[0]
                    selected_matches.append(best_match)
                
                # 再次处理冲突
                matches_by_unified = {}
                for match in selected_matches:
                    unified_col = match['matched']
                    if unified_col not in matches_by_unified:
                        matches_by_unified[unified_col] = []
                    matches_by_unified[unified_col].append(match)
                
                final_selected_matches = []
                for unified_col, matches in matches_by_unified.items():
                    sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                    best_match = sorted_matches[0]
                    final_selected_matches.append(best_match)
                
                selected_matches = final_selected_matches
                
                # 应用选定的匹配
                for match in selected_matches:
                    original_col = match['original']
                    unified_col = match['matched']
                    if unified_col in mapped_df.columns:
                        mapped_df[unified_col] = df_temp[original_col]
                        matched_cols.append(match)
                
                # 排除全NA列
                mapped_df = mapped_df.dropna(axis=1, how='all')
                
                # 将NaN替换为空字符串
                mapped_df = mapped_df.fillna('').infer_objects()
                
                return mapped_df if not mapped_df.empty else None
        
        # 创建临时DataFrame进行映射
        temp_df = pd.DataFrame([best_header.tolist()], columns=best_header.tolist())
        temp_result = temp_map(temp_df)
        
        if temp_result is not None:
            # 检查是否缺少户主名称
            mapped_headers = list(temp_result.columns)
            has_missing_account_name = '户主名称' not in mapped_headers
        
        return has_missing_account_name, current_best_header_score, has_other_header_rows
    
    def _check_header_for_account_name(self, header_df: pd.DataFrame) -> tuple:
        """
        检查表头是否缺少户主名称
        
        Args:
            header_df: 表头数据
            
        Returns:
            (是否缺少户主名称, 最佳表头行评分, 是否存在其他表头行评分大于等于最佳表头行评分)
        """
        # 寻找最佳表头
        current_best_header_row = 0
        current_best_header_score = 0
        
        for i in range(min(50, len(header_df))):
            header_candidate = header_df.iloc[i]
            standardized_headers = [standardize_header(str(h)) for h in header_candidate]
            score = self._calculate_header_score(standardized_headers)
            
            if score > current_best_header_score:
                current_best_header_score = score
                current_best_header_row = i
        
        # 检查是否存在其他表头行评分大于等于最佳表头行评分
        has_other_header_rows = False
        for i in range(min(50, len(header_df))):
            if i == current_best_header_row:
                continue
            header_candidate = header_df.iloc[i]
            standardized_headers = [standardize_header(str(h)) for h in header_candidate]
            score = self._calculate_header_score(standardized_headers)
            if score >= current_best_header_score:
                has_other_header_rows = True
                break
        
        # 检查最佳表头是否缺少户主名称
        has_missing_account_name = True
        if current_best_header_row < len(header_df):
            best_header = header_df.iloc[current_best_header_row].apply(standardize_header)
            # 创建临时映射函数，检查是否缺少户主名称
            def temp_map(df_temp):
                # 表头标准化
                df_temp.columns = [standardize_header(col) for col in df_temp.columns]
                standardized_headers = list(df_temp.columns)
                
                # 映射到统一表头
                mapped_df = pd.DataFrame(columns=self.unified_headers)
                matched_cols = []
                
                # 收集所有可能的匹配
                all_matches = []
                
                # 1. 收集所有精确匹配
                for col in standardized_headers:
                    if col in self.header_mapping:
                        unified_col = self.header_mapping[col]
                        mapping_items = list(self.header_mapping.items())
                        position = next((i for i, (k, v) in enumerate(mapping_items) if k == col), len(mapping_items))
                        position_weight = max(0, 10 - position * 0.1)
                        weight = 20 + position_weight
                        
                        all_matches.append({
                            'original': col,
                            'matched': unified_col,
                            'method': '精确匹配',
                            'weight': weight,
                            'position': position
                        })
                
                # 2. 收集所有关键词匹配
                for col in standardized_headers:
                    if any(m['original'] == col and m['method'] == '精确匹配' for m in all_matches):
                        continue
                    
                    for unified_col, keywords in self.keyword_mapping.items():
                        for i, keyword in enumerate(keywords):
                            if unified_col in self.exclusion_rules:
                                exclusion_keywords = self.exclusion_rules[unified_col]
                                if any(exclusion_keyword in col for exclusion_keyword in exclusion_keywords):
                                    continue
                            
                            score = 0
                            if col == keyword:
                                score = 10
                            elif col.startswith(keyword) or col.endswith(keyword):
                                score = 8
                            elif keyword in col:
                                keyword_chars = set(keyword)
                                col_chars = set(col)
                                if keyword_chars:
                                    common_chars = keyword_chars.intersection(col_chars)
                                    coverage_ratio = len(common_chars) / len(keyword_chars)
                                    if coverage_ratio > 0.5:
                                        score = 5
                            
                            if score > 0:
                                keyword_position_weight = max(0, 5 - i * 0.5)
                                weight = score + keyword_position_weight
                                
                                all_matches.append({
                                    'original': col,
                                    'matched': unified_col,
                                    'method': '关键词匹配',
                                    'keyword': keyword,
                                    'weight': weight,
                                    'score': score,
                                    'keyword_position': i
                                })
                
                # 处理匹配冲突
                matches_by_original = {}
                for match in all_matches:
                    original_col = match['original']
                    if original_col not in matches_by_original:
                        matches_by_original[original_col] = []
                    matches_by_original[original_col].append(match)
                
                selected_matches = []
                for original_col, matches in matches_by_original.items():
                    sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                    best_match = sorted_matches[0]
                    selected_matches.append(best_match)
                
                # 再次处理冲突
                matches_by_unified = {}
                for match in selected_matches:
                    unified_col = match['matched']
                    if unified_col not in matches_by_unified:
                        matches_by_unified[unified_col] = []
                    matches_by_unified[unified_col].append(match)
                
                final_selected_matches = []
                for unified_col, matches in matches_by_unified.items():
                    sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                    best_match = sorted_matches[0]
                    final_selected_matches.append(best_match)
                
                selected_matches = final_selected_matches
                
                # 应用选定的匹配
                for match in selected_matches:
                    original_col = match['original']
                    unified_col = match['matched']
                    if unified_col in mapped_df.columns:
                        mapped_df[unified_col] = df_temp[original_col]
                        matched_cols.append(match)
                
                # 排除全NA列
                mapped_df = mapped_df.dropna(axis=1, how='all')
                
                # 将NaN替换为空字符串
                mapped_df = mapped_df.fillna('').infer_objects()
                
                return mapped_df if not mapped_df.empty else None
        
        # 创建临时DataFrame进行映射
        temp_df = pd.DataFrame([best_header.tolist()], columns=best_header.tolist())
        temp_result = temp_map(temp_df)
        
        if temp_result is not None:
            # 检查是否缺少户主名称
            mapped_headers = list(temp_result.columns)
            has_missing_account_name = '户主名称' not in mapped_headers
        
        return has_missing_account_name, current_best_header_score, has_other_header_rows
    
    def _need_multi_account_processing(self, file_path: str) -> tuple:
        """
        检查大文件是否需要多账户处理
        
        Args:
            file_path: 文件路径
            
        Returns:
            (是否需要多账户处理, 最优表头行评分, 是否需要无户名单账户处理)
        """
        try:
            from src.utils.file_utils import is_excel_file, is_text_file
            
            best_header_score = 0
            need_single_account_no_name = False
            
            if is_excel_file(file_path):
                # 读取前100行寻找表头
                excel_file = pd.ExcelFile(file_path)
                
                # 检查所有sheet
                for sheet_name in excel_file.sheet_names:
                    self.logger.info(f"【大文件处理】检查sheet：{sheet_name}")
                    header_df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        nrows=100
                    )
                    
                    # 检查表头是否缺少户主名称
                    has_missing_account_name, current_best_header_score, has_other_header_rows = self._check_header_for_account_name(header_df)
                    
                    # 更新全局最佳表头得分
                    if current_best_header_score > best_header_score:
                        best_header_score = current_best_header_score
                    
                    if has_missing_account_name and has_other_header_rows:
                        self.logger.info(f"【大文件处理】检测到{sheet_name}表头缺少户主名称且存在其他表头行，需要多账户处理")
                        return True, best_header_score, False
                    elif has_missing_account_name and not has_other_header_rows:
                        self.logger.info(f"【大文件处理】检测到{sheet_name}表头缺少户主名称且不存在其他表头行，需要无户名单账户处理")
                        need_single_account_no_name = True
            
            elif is_text_file(file_path):
                # 类似Excel文件的处理逻辑
                encoding = detect_file_encoding(file_path, self.config)
                delimiter = detect_file_delimiter(file_path, encoding, self.config)
                
                # 读取前100行寻找表头
                header_df = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    sep=delimiter,
                    header=None,
                    nrows=100,
                    skip_blank_lines=True,
                    on_bad_lines='skip'
                )
                
                # 检查表头是否缺少户主名称
                has_missing_account_name, current_best_header_score, has_other_header_rows = self._check_header_for_account_name(header_df)
                
                # 更新全局最佳表头得分
                best_header_score = current_best_header_score
                
                if has_missing_account_name and has_other_header_rows:
                    self.logger.info(f"【大文件处理】检测到表头缺少户主名称且存在其他表头行，需要多账户处理")
                    return True, best_header_score, False
                elif has_missing_account_name and not has_other_header_rows:
                    self.logger.info(f"【大文件处理】检测到表头缺少户主名称且不存在其他表头行，需要无户名单账户处理")
                    need_single_account_no_name = True
            
            return False, best_header_score, need_single_account_no_name
        except Exception as e:
            self.logger.error(f"【多账户处理检测异常】{file_path}：{str(e)}", exc_info=True)
            return False, 0, False
    
    def _process_large_single_account_no_name(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        处理大型无户名单账户文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件无户名单账户处理】开始处理：{file_name}")
        
        try:
            import gc
            from src.utils.file_utils import is_excel_file, is_text_file
            
            # 智能线程池管理
            def get_optimal_thread_count():
                try:
                    import psutil
                    cpu_count = psutil.cpu_count(logical=True)
                    memory = psutil.virtual_memory()
                    # 根据内存和CPU数量动态调整线程数
                    if memory.percent > 70:
                        return max(1, cpu_count // 2)
                    else:
                        return max(1, cpu_count)
                except:
                    return 4  # 默认线程数
            
            optimal_threads = get_optimal_thread_count()
            self.logger.info(f"【并行处理】使用最优线程数：{optimal_threads}")
            
            # 全篇读取文件，注意内存使用
            if is_excel_file(file_path):
                # 读取Excel文件
                excel_file = pd.ExcelFile(file_path)
                
                # 优化数据流：直接处理并合并，减少中间存储
                combined_df = None
                batch_size = 5  # 每批处理的sheet数量
                
                for i, sheet_name in enumerate(excel_file.sheet_names):
                    self.logger.info(f"【大文件无户名单账户处理】处理Excel sheet：{sheet_name}")
                    # 全篇读取sheet数据
                    try:
                        # 使用低内存模式读取
                        df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                        self.logger.info(f"【大文件无户名单账户处理】读取完成，共{len(df_raw)}行数据")
                        
                        # 处理无户名单账户数据
                        sheet_result = self.multi_account_handler._process_single_account_no_name(df_raw)
                        
                        # 释放内存
                        del df_raw
                        gc.collect()
                        
                        if sheet_result is not None:
                            # 处理数据
                            processed_result = self._process_file_data(sheet_result, f"{file_name} - {sheet_name}")
                            if processed_result is not None and not processed_result.empty:
                                # 直接合并到结果中，减少中间存储
                                if combined_df is None:
                                    combined_df = processed_result
                                else:
                                    # 使用inplace=True减少数据复制
                                    combined_df = pd.concat([combined_df, processed_result], ignore_index=True)
                                # 释放内存
                                del processed_result
                                gc.collect()
                    except MemoryError:
                        self.logger.error(f"【大文件无户名单账户处理】内存溢出，尝试分块处理：{sheet_name}")
                        # 尝试分块处理
                        try:
                            # 先读取全文识别最佳表头行
                            full_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                            # 寻找最佳表头行
                            best_header_row, best_header_score = self.multi_account_handler._find_best_header_row(full_df)
                            
                            if best_header_row > 0:
                                # 分块处理
                                chunk_size = 10000
                                for i in range(0, len(full_df), chunk_size):
                                    end_row = min(i + chunk_size, len(full_df))
                                    # 直接使用切片，避免copy()
                                    chunk_df = full_df.iloc[i:end_row]
                                    chunk_result = self.multi_account_handler._process_single_account_no_name(chunk_df)
                                    if chunk_result is not None:
                                        processed_chunk = self._process_file_data(chunk_result, f"{file_name} - {sheet_name} - 块")
                                        if processed_chunk is not None and not processed_chunk.empty:
                                            # 直接合并到结果中
                                            if combined_df is None:
                                                combined_df = processed_chunk
                                            else:
                                                combined_df = pd.concat([combined_df, processed_chunk], ignore_index=True)
                                            # 释放内存
                                            del processed_chunk
                                            gc.collect()
                            
                            # 释放full_df内存
                            del full_df
                            gc.collect()
                        except Exception as e:
                            self.logger.error(f"【大文件无户名单账户处理】分块处理异常：{str(e)}")
                
                if combined_df is not None:
                    self.logger.info(f"【大文件无户名单账户处理】处理完成，共{len(combined_df)}行数据")
                    return combined_df
                else:
                    return None
            
            elif is_text_file(file_path):
                # 读取文本文件
                encoding = detect_file_encoding(file_path, self.config)
                delimiter = detect_file_delimiter(file_path, encoding, self.config)
                
                try:
                    # 全篇读取文件
                    df_raw = pd.read_csv(
                        file_path,
                        encoding=encoding,
                        sep=delimiter,
                        header=None,
                        low_memory=False,
                        skip_blank_lines=True,
                        on_bad_lines='skip'
                    )
                    self.logger.info(f"【大文件无户名单账户处理】读取完成，共{len(df_raw)}行数据")
                    
                    # 处理无户名单账户数据
                    single_account_df = self.multi_account_handler._process_single_account_no_name(df_raw)
                    
                    # 释放内存
                    del df_raw
                    gc.collect()
                    
                    if single_account_df is not None:
                        # 处理数据
                        processed_result = self._process_file_data(single_account_df, file_name)
                        if processed_result is not None and not processed_result.empty:
                            self.logger.info(f"【大文件无户名单账户处理】处理完成，共{len(processed_result)}行数据")
                            return processed_result
                except MemoryError:
                    self.logger.error(f"【大文件无户名单账户处理】内存溢出，尝试分块处理")
                    # 尝试分块处理
                    try:
                        # 先读取全文寻找最佳表头行
                        full_df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            sep=delimiter,
                            header=None,
                            low_memory=False,
                            skip_blank_lines=True,
                            on_bad_lines='skip'
                        )
                        
                        # 寻找最佳表头行
                        best_header_row, best_header_score = self.multi_account_handler._find_best_header_row(full_df)
                        
                        if best_header_row > 0:
                            # 分块处理
                            chunk_size = 10000
                            combined_df = None
                            for i in range(0, len(full_df), chunk_size):
                                end_row = min(i + chunk_size, len(full_df))
                                chunk_df = pd.read_csv(
                                    file_path,
                                    encoding=encoding,
                                    sep=delimiter,
                                    header=None,
                                    skiprows=i,
                                    nrows=end_row-i,
                                    skip_blank_lines=True,
                                    on_bad_lines='skip'
                                )
                                chunk_result = self.multi_account_handler._process_single_account_no_name(chunk_df)
                                if chunk_result is not None:
                                    processed_chunk = self._process_file_data(chunk_result, f"{file_name} - 块")
                                    if processed_chunk is not None and not processed_chunk.empty:
                                        # 直接合并到结果中
                                        if combined_df is None:
                                            combined_df = processed_chunk
                                        else:
                                            combined_df = pd.concat([combined_df, processed_chunk], ignore_index=True)
                                        # 释放内存
                                        del processed_chunk
                                        gc.collect()
                        
                        # 释放full_df内存
                        del full_df
                        gc.collect()
                        
                        if combined_df is not None:
                            self.logger.info(f"【大文件无户名单账户处理】处理完成，共{len(combined_df)}行数据")
                            return combined_df
                    except Exception as e:
                        self.logger.error(f"【大文件无户名单账户处理】分块处理异常：{str(e)}")
                
                return None
            
            return None
        except Exception as e:
            self.logger.error(f"【大文件无户名单账户处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_large_multi_account_file(self, file_path: str, best_header_score: float = 0) -> Optional[pd.DataFrame]:
        """
        处理大型多账户文件
        
        Args:
            file_path: 文件路径
            best_header_score: 最优表头行评分
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件多账户处理】开始处理：{file_name}")
        self.logger.info(f"【大文件多账户处理】最优表头行评分：{best_header_score}")
        
        # 添加内存使用监控
        def log_memory_usage(stage):
            try:
                import psutil
                process = psutil.Process()
                memory_info = process.memory_info()

            except:
                pass
        
        log_memory_usage("开始处理")
        
        try:
            import gc
            from concurrent.futures import ThreadPoolExecutor, as_completed
            from src.utils.file_utils import is_excel_file, is_text_file
            
            # 智能线程池管理
            def get_optimal_thread_count():
                try:
                    import psutil
                    cpu_count = psutil.cpu_count(logical=True)
                    memory = psutil.virtual_memory()
                    # 根据内存和CPU数量动态调整线程数
                    if memory.percent > 70:
                        return max(1, cpu_count // 2)
                    else:
                        return max(1, cpu_count)
                except:
                    return 4  # 默认线程数
            
            optimal_threads = get_optimal_thread_count()
            self.logger.info(f"【并行处理】使用最优线程数：{optimal_threads}")
            
            # 全篇读取文件，注意内存使用
            if is_excel_file(file_path):
                # 读取Excel文件
                excel_file = pd.ExcelFile(file_path)
                
                # 优化数据流：直接处理并合并，减少中间存储
                combined_df = None
                batch_size = 5  # 每批处理的sheet数量
                
                for i, sheet_name in enumerate(excel_file.sheet_names):
                    self.logger.info(f"【大文件多账户处理】处理Excel sheet：{sheet_name}")
                    log_memory_usage(f"处理sheet {sheet_name}")
                    # 全篇读取sheet数据
                    try:
                        # 使用低内存模式读取
                        df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                        self.logger.info(f"【大文件多账户处理】读取完成，共{len(df_raw)}行数据")
                        log_memory_usage(f"读取sheet {sheet_name}完成")
                        
                        # 处理多账户数据
                        sheet_result = self.multi_account_handler.process_multi_account_sheet(df_raw, f"{file_name} - {sheet_name}", best_header_score)
                        
                        # 释放内存
                        del df_raw
                        gc.collect()
                        log_memory_usage(f"释放sheet {sheet_name}内存")
                        
                        if sheet_result is not None:
                            # 处理数据
                            processed_result = self._process_file_data(sheet_result, f"{file_name} - {sheet_name}")
                            if processed_result is not None and not processed_result.empty:
                                # 直接合并到结果中，减少中间存储
                                if combined_df is None:
                                    combined_df = processed_result
                                else:
                                    # 使用inplace=True减少数据复制
                                    combined_df = pd.concat([combined_df, processed_result], ignore_index=True)
                                # 释放内存
                                del processed_result
                                gc.collect()
                                log_memory_usage(f"处理sheet {sheet_name}完成")
                    except MemoryError:
                        self.logger.error(f"【大文件多账户处理】内存溢出，尝试分块处理：{sheet_name}")
                        # 尝试分块处理
                        try:
                            # 先读取全文识别分隔行
                            full_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                            # 识别分隔行
                            separator_candidates = []
                            for i in range(len(full_df)):
                                score = self.multi_account_handler._calculate_separator_score(full_df.iloc[i])
                                if score > 0:
                                    separator_candidates.append((i, score))
                            
                            if separator_candidates:
                                # 选择最佳分隔行
                                max_score = max(score for _, score in separator_candidates)
                                best_separators = [row for row, score in separator_candidates if score == max_score]
                                best_separators.sort()
                                
                                # 分块处理
                                start_row = 0
                                for separator_row in best_separators:
                                    end_row = separator_row
                                    if start_row < end_row:
                                        # 直接使用切片，避免copy()
                                        chunk_df = full_df.iloc[start_row:end_row]
                                        chunk_result = self.multi_account_handler.process_multi_account_sheet(chunk_df, f"{file_name} - {sheet_name} - 块1", best_header_score)
                                        if chunk_result is not None:
                                            processed_chunk = self._process_file_data(chunk_result, f"{file_name} - {sheet_name} - 块1")
                                            if processed_chunk is not None and not processed_chunk.empty:
                                                # 直接合并到结果中
                                                if combined_df is None:
                                                    combined_df = processed_chunk
                                                else:
                                                    combined_df = pd.concat([combined_df, processed_chunk], ignore_index=True)
                                                # 释放内存
                                                del processed_chunk
                                                gc.collect()
                                    start_row = separator_row
                                
                                # 处理最后一块
                                if start_row < len(full_df):
                                    # 直接使用切片，避免copy()
                                    chunk_df = full_df.iloc[start_row:]
                                    chunk_result = self.multi_account_handler.process_multi_account_sheet(chunk_df, f"{file_name} - {sheet_name} - 最后块", best_header_score)
                                    if chunk_result is not None:
                                        processed_chunk = self._process_file_data(chunk_result, f"{file_name} - {sheet_name} - 最后块")
                                        if processed_chunk is not None and not processed_chunk.empty:
                                            # 直接合并到结果中
                                            if combined_df is None:
                                                combined_df = processed_chunk
                                            else:
                                                combined_df = pd.concat([combined_df, processed_chunk], ignore_index=True)
                                            # 释放内存
                                            del processed_chunk
                                            gc.collect()
                                
                                # 释放full_df内存
                                del full_df
                                gc.collect()
                        except Exception as e:
                            self.logger.error(f"【大文件多账户处理】分块处理异常：{str(e)}")
                
                if combined_df is not None:
                    self.logger.info(f"【大文件多账户处理】处理完成，共{len(combined_df)}行数据")
                    log_memory_usage("处理完成")
                    return combined_df
                else:
                    return None
            
            elif is_text_file(file_path):
                # 读取文本文件
                encoding = detect_file_encoding(file_path, self.config)
                delimiter = detect_file_delimiter(file_path, encoding, self.config)
                
                try:
                    # 全篇读取文件
                    df_raw = pd.read_csv(
                        file_path,
                        encoding=encoding,
                        sep=delimiter,
                        header=None,
                        low_memory=False,
                        skip_blank_lines=True,
                        on_bad_lines='skip'
                    )
                    self.logger.info(f"【大文件多账户处理】读取完成，共{len(df_raw)}行数据")
                    log_memory_usage("读取文件完成")
                    
                    # 处理多账户数据
                    multi_account_df = self.multi_account_handler.process_multi_account_sheet(df_raw, file_name, best_header_score)
                    
                    # 释放内存
                    del df_raw
                    gc.collect()
                    log_memory_usage("释放文件内存")
                    
                    if multi_account_df is not None:
                        # 处理数据
                        processed_result = self._process_file_data(multi_account_df, file_name)
                        if processed_result is not None and not processed_result.empty:
                            self.logger.info(f"【大文件多账户处理】处理完成，共{len(processed_result)}行数据")
                            log_memory_usage("处理完成")
                            return processed_result
                except MemoryError:
                    self.logger.error(f"【大文件多账户处理】内存溢出，尝试分块处理")
                    # 尝试分块处理
                    try:
                        # 先读取全文识别分隔行
                        full_df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            sep=delimiter,
                            header=None,
                            low_memory=False,
                            skip_blank_lines=True,
                            on_bad_lines='skip'
                        )
                        
                        # 识别分隔行
                        separator_candidates = []
                        for i in range(len(full_df)):
                            score = self.multi_account_handler._calculate_separator_score(full_df.iloc[i])
                            if score > 0:
                                separator_candidates.append((i, score))
                        
                        if separator_candidates:
                            # 选择最佳分隔行
                            max_score = max(score for _, score in separator_candidates)
                            best_separators = [row for row, score in separator_candidates if score == max_score]
                            best_separators.sort()
                            
                            # 优化数据流：直接处理并合并
                            combined_df = None
                            start_row = 0
                            
                            for separator_row in best_separators:
                                end_row = separator_row
                                if start_row < end_row:
                                    chunk_df = pd.read_csv(
                                        file_path,
                                        encoding=encoding,
                                        sep=delimiter,
                                        header=None,
                                        skiprows=start_row,
                                        nrows=end_row-start_row,
                                        skip_blank_lines=True,
                                        on_bad_lines='skip'
                                    )
                                    chunk_result = self.multi_account_handler.process_multi_account_sheet(chunk_df, f"{file_name} - 块", best_header_score)
                                    if chunk_result is not None:
                                        processed_chunk = self._process_file_data(chunk_result, f"{file_name} - 块")
                                        if processed_chunk is not None and not processed_chunk.empty:
                                            # 直接合并到结果中
                                            if combined_df is None:
                                                combined_df = processed_chunk
                                            else:
                                                combined_df = pd.concat([combined_df, processed_chunk], ignore_index=True)
                                            # 释放内存
                                            del processed_chunk
                                            gc.collect()
                                start_row = separator_row
                            
                            # 处理最后一块
                            if start_row > 0:
                                chunk_df = pd.read_csv(
                                    file_path,
                                    encoding=encoding,
                                    sep=delimiter,
                                    header=None,
                                    skiprows=start_row,
                                    skip_blank_lines=True,
                                    on_bad_lines='skip'
                                )
                                chunk_result = self.multi_account_handler.process_multi_account_sheet(chunk_df, f"{file_name} - 最后块", best_header_score)
                                if chunk_result is not None:
                                    processed_chunk = self._process_file_data(chunk_result, f"{file_name} - 最后块")
                                    if processed_chunk is not None and not processed_chunk.empty:
                                        # 直接合并到结果中
                                        if combined_df is None:
                                            combined_df = processed_chunk
                                        else:
                                            combined_df = pd.concat([combined_df, processed_chunk], ignore_index=True)
                                        # 释放内存
                                        del processed_chunk
                                        gc.collect()
                            
                            # 释放full_df内存
                            del full_df
                            gc.collect()
                            log_memory_usage("分块处理完成")
                            
                            if combined_df is not None:
                                self.logger.info(f"【大文件多账户处理】处理完成，共{len(combined_df)}行数据")
                                log_memory_usage("处理完成")
                                return combined_df
                    except Exception as e:
                        self.logger.error(f"【大文件多账户处理】分块处理异常：{str(e)}")
                
                return None
            
            return None
        except Exception as e:
            self.logger.error(f"【大文件多账户处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_large_excel_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        分块处理大型Excel文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件处理】开始处理大型Excel文件：{file_name}")
        
        try:
            # 预处理Excel文件
            processed_file_path = self._preprocess_excel_file(file_path)
            try:
                # 使用pandas的read_excel分块读取
                excel_file = pd.ExcelFile(processed_file_path)
                all_sheets_data = []
                
                for sheet_name in excel_file.sheet_names:
                    self.logger.info(f"【大文件处理】处理Excel sheet：{sheet_name}")
                    
                    # 首先读取前100行寻找表头
                    header_df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        nrows=100
                    )
                    
                    # 寻找最佳表头
                    best_header_row = 0
                    best_header_score = 0
                    
                    for i in range(min(50, len(header_df))):
                        header_candidate = header_df.iloc[i]
                        standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                        score = self._calculate_header_score(standardized_headers)
                        
                        if score > best_header_score:
                            best_header_score = score
                            best_header_row = i
                    
                    # 提取表头
                    header_row = header_df.iloc[best_header_row]
                    headers = [standardize_header(str(h)) for h in header_row]
                    
                    # 读取整个sheet数据
                    sheet_df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        skiprows=best_header_row + 1
                    )
                    
                    # 处理列数不一致的情况
                    if len(sheet_df.columns) > len(headers):
                        sheet_df = sheet_df.iloc[:, :len(headers)]
                        self.logger.debug(f"【大文件处理】{file_name} 列数多于表头，截断为{len(headers)}列")
                    elif len(sheet_df.columns) < len(headers):
                        for j in range(len(sheet_df.columns), len(headers)):
                            sheet_df[j] = None
                        self.logger.debug(f"【大文件处理】{file_name} 列数少于表头，添加到{len(headers)}列")
                    
                    # 设置表头
                    sheet_df.columns = headers
                    
                    # 分块处理数据
                    chunks = []
                    total_rows = 0
                    
                    for i in range(0, len(sheet_df), self.chunk_size):
                        end = min(i + self.chunk_size, len(sheet_df))
                        chunk = sheet_df.iloc[i:end]
                        self.logger.info(f"【大文件处理】{file_name} - {sheet_name} 处理第{i//self.chunk_size+1}块，共{len(chunk)}行")
                        
                        # 处理数据
                        processed_chunk = self._process_file_data(chunk, f"{file_name} - {sheet_name} - 块{i//self.chunk_size+1}")
                        if processed_chunk is not None and not processed_chunk.empty:
                            chunks.append(processed_chunk)
                            total_rows += len(processed_chunk)
                    
                    if chunks:
                        sheet_df = pd.concat(chunks, ignore_index=True)
                        all_sheets_data.append(sheet_df)
                        self.logger.info(f"【大文件处理】{file_name} - {sheet_name} 处理完成，共{total_rows}行数据")
                
                if all_sheets_data:
                    combined_df = pd.concat(all_sheets_data, ignore_index=True)
                    self.logger.info(f"【大文件处理】{file_name} 所有sheet处理完成，共{len(combined_df)}行数据")
                    return combined_df
                else:
                    self.logger.warning(f"【大文件处理】{file_name} 没有有效数据")
                    return None
            finally:
                # 清理临时文件
                if processed_file_path != file_path:
                    if os.path.exists(processed_file_path):
                        try:
                            os.remove(processed_file_path)
                        except Exception as e:
                            self.logger.debug(f"【临时文件清理】无法删除文件 {processed_file_path}：{str(e)}")
        except Exception as e:
            self.logger.error(f"【大文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    

    
    def _process_large_text_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        分块处理大型文本文件（CSV/TXT/HTML）
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件处理】开始处理大型文本文件：{file_name}")
        
        try:
            # 导入文件类型检测函数
            from src.utils.file_utils import detect_file_type, get_file_extension_from_mime, magic_available
            
            # 检测文件编码（使用缓存）
            if file_path not in self.encoding_cache:
                self.encoding_cache[file_path] = detect_file_encoding(file_path, self.config)
            encoding = self.encoding_cache[file_path]
            # 更新访问时间
            self._update_cache_access("encoding", file_path)
            
            # 检测文件类型
            file_type = detect_file_type(file_path)
            file_ext = get_file_extension_from_mime(file_type)
            
            # 判断是否为CSV文件
            is_csv = False
            if magic_available:
                is_csv = file_type in ['text/csv', 'application/csv']
            else:
                is_csv = file_path.lower().endswith('.csv')
            
            # 判断是否为HTML文件
            is_html = False
            if magic_available:
                is_html = file_type == 'text/html'
            else:
                is_html = file_path.lower().endswith(('.html', '.htm'))
            
            if is_html:
                # 处理大型HTML文件
                return self._process_large_html_file(file_path, encoding)
            elif is_csv:
                # 处理大型CSV文件
                return self._process_large_csv_file(file_path, encoding)
            else:
                # 处理大型TXT文件
                return self._process_large_txt_file(file_path, encoding)
        except Exception as e:
            self.logger.error(f"【大文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_large_csv_file(self, file_path: str, encoding: str) -> Optional[pd.DataFrame]:
        """
        分块处理大型CSV文件
        
        Args:
            file_path: 文件路径
            encoding: 文件编码
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件处理】开始处理大型CSV文件：{file_name}")
        
        try:
            # 检测分隔符（使用缓存）
            if file_path not in self.delimiter_cache:
                from src.utils import detect_file_delimiter
                self.delimiter_cache[file_path] = detect_file_delimiter(file_path, encoding, self.config)
            delimiter = self.delimiter_cache[file_path]
            # 更新访问时间
            self._update_cache_access("delimiter", file_path)
            
            self.logger.info(f"【大文件处理】{file_name} 分块大小：{self.chunk_size}行")
            
            # 分块读取文件
            chunks = []
            total_rows = 0
            
            # 首先扫描前100行寻找表头
            header_df = pd.read_csv(
                file_path,
                encoding=encoding,
                sep=delimiter,
                header=None,
                nrows=100,
                skip_blank_lines=True,
                on_bad_lines='skip'
            )
            
            # 寻找最佳表头
            best_header_row = 0
            best_header_score = 0
            
            for i in range(min(50, len(header_df))):
                header_candidate = header_df.iloc[i]
                standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                score = self._calculate_header_score(standardized_headers)
                
                if score > best_header_score:
                    best_header_score = score
                    best_header_row = i
            
            # 提取表头
            header_row = header_df.iloc[best_header_row]
            headers = [standardize_header(str(h)) for h in header_row]
            
            # 分块读取数据，使用正确的 skiprows 参数
            chunk_iter = pd.read_csv(
                file_path,
                encoding=encoding,
                sep=delimiter,
                header=None,
                chunksize=self.chunk_size,
                skip_blank_lines=True,
                on_bad_lines='skip',
                skiprows=best_header_row + 1
            )
            
            for i, chunk in enumerate(chunk_iter):
                self.logger.info(f"【大文件处理】{file_name} 处理第{i+1}块，共{len(chunk)}行")
                
                # 处理列数不一致的情况
                if len(chunk.columns) > len(headers):
                    # 列数多于表头，截断
                    chunk = chunk.iloc[:, :len(headers)]
                    self.logger.debug(f"【大文件处理】{file_name} 块{i+1} 列数多于表头，截断为{len(headers)}列")
                elif len(chunk.columns) < len(headers):
                    # 列数少于表头，添加空列
                    for j in range(len(chunk.columns), len(headers)):
                        chunk[j] = None
                    self.logger.debug(f"【大文件处理】{file_name} 块{i+1} 列数少于表头，添加到{len(headers)}列")
                
                # 设置表头
                chunk.columns = headers
                
                # 处理数据
                processed_chunk = self._process_file_data(chunk, f"{file_name} - 块{i+1}")
                if processed_chunk is not None and not processed_chunk.empty:
                    chunks.append(processed_chunk)
                    total_rows += len(processed_chunk)
            
            if chunks:
                # 合并所有块
                combined_df = pd.concat(chunks, ignore_index=True)
                self.logger.info(f"【大文件处理】{file_name} 处理完成，共{total_rows}行数据")
                return combined_df
            else:
                self.logger.warning(f"【大文件处理】{file_name} 没有有效数据")
                return None
        except Exception as e:
            self.logger.error(f"【大文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_large_txt_file(self, file_path: str, encoding: str) -> Optional[pd.DataFrame]:
        """
        分块处理大型TXT文件
        
        Args:
            file_path: 文件路径
            encoding: 文件编码
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件处理】开始处理大型TXT文件：{file_name}")
        
        try:
            # 预处理TXT文件
            from src.utils import preprocess_txt_file
            actual_file_path = preprocess_txt_file(file_path, encoding, self.config)
            
            # 定义常见分隔符列表
            common_delimiters = [',', '\t', ';', '|', ' ']
            
            # 检测分隔符（使用缓存）
            if actual_file_path not in self.delimiter_cache:
                from src.utils import detect_file_delimiter
                self.delimiter_cache[actual_file_path] = detect_file_delimiter(actual_file_path, encoding, self.config)
            delimiter = self.delimiter_cache[actual_file_path]
            # 更新访问时间
            self._update_cache_access("delimiter", actual_file_path)
            
            self.logger.info(f"【大文件处理】{file_name} 分块大小：{self.chunk_size}行")
            
            # 分块读取文件
            chunks = []
            total_rows = 0
            
            # 首先扫描前100行寻找表头
            header_df = pd.read_csv(
                actual_file_path,
                encoding=encoding,
                sep=delimiter,
                header=None,
                nrows=100,
                skip_blank_lines=True,
                on_bad_lines='skip'
            )
            
            # 寻找最佳表头
            best_header_row = 0
            best_header_score = 0
            
            for i in range(min(50, len(header_df))):
                header_candidate = header_df.iloc[i]
                standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                score = self._calculate_header_score(standardized_headers)
                
                if score > best_header_score:
                    best_header_score = score
                    best_header_row = i
            
            # 提取表头
            header_row = header_df.iloc[best_header_row]
            headers = [standardize_header(str(h)) for h in header_row]
            
            # 分块读取数据
            chunk_iter = pd.read_csv(
                actual_file_path,
                encoding=encoding,
                sep=delimiter,
                header=None,
                chunksize=self.chunk_size,
                skip_blank_lines=True,
                on_bad_lines='skip',
                skiprows=best_header_row + 1
            )
            
            for i, chunk in enumerate(chunk_iter):
                self.logger.info(f"【大文件处理】{file_name} 处理第{i+1}块，共{len(chunk)}行")
                
                # 处理列数不一致的情况
                if len(chunk.columns) > len(headers):
                    chunk = chunk.iloc[:, :len(headers)]
                    self.logger.debug(f"【大文件处理】{file_name} 块{i+1} 列数多于表头，截断为{len(headers)}列")
                elif len(chunk.columns) < len(headers):
                    for j in range(len(chunk.columns), len(headers)):
                        chunk[j] = None
                    self.logger.debug(f"【大文件处理】{file_name} 块{i+1} 列数少于表头，添加到{len(headers)}列")
                
                # 设置表头
                chunk.columns = headers
                
                # 处理数据
                processed_chunk = self._process_file_data(chunk, f"{file_name} - 块{i+1}")
                if processed_chunk is not None and not processed_chunk.empty:
                    chunks.append(processed_chunk)
                    total_rows += len(processed_chunk)
            
            if chunks:
                combined_df = pd.concat(chunks, ignore_index=True)
                self.logger.info(f"【大文件处理】{file_name} 处理完成，共{total_rows}行数据")
                return combined_df
            else:
                self.logger.warning(f"【大文件处理】{file_name} 没有有效数据")
                return None
        except Exception as e:
            self.logger.error(f"【大文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_large_html_file(self, file_path: str, encoding: str) -> Optional[pd.DataFrame]:
        """
        处理大型HTML文件
        
        Args:
            file_path: 文件路径
            encoding: 文件编码
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        self.logger.info(f"【大文件处理】开始处理大型HTML文件：{file_name}")
        
        try:
            # 使用BeautifulSoup解析HTML
            from bs4 import BeautifulSoup
            
            with open(file_path, 'r', encoding=encoding) as f:
                soup = BeautifulSoup(f, 'lxml')
            
            # 检测是否为农业银行HTML文件
            if self._is_abc_bank_html(soup):
                # 特殊处理农业银行HTML文件
                return self._process_abc_bank_html(file_path, encoding)
            
            # 提取所有表格
            tables = soup.find_all('table')
            
            if not tables:
                self.logger.error(f"【HTML处理】文件 {file_name} 没有找到表格")
                return None
            
            # 表格评分和选择
            best_table = None
            best_score = 0
            
            for table in tables:
                try:
                    # 转换为DataFrame
                    from io import StringIO
                    table_html = str(table)
                    table_dfs = pd.read_html(StringIO(table_html), header=None)
                    if not table_dfs:
                        continue
                    
                    table_df = table_dfs[0]
                    
                    # 计算表格得分
                    row_score = min(len(table_df), 100) / 100
                    col_count = len(table_df.columns)
                    if 10 <= col_count <= 20:
                        col_score = 1.0
                    elif 5 <= col_count < 10 or 20 < col_count <= 30:
                        col_score = 0.7
                    else:
                        col_score = 0.3
                    non_empty_cells = table_df.notna().sum().sum()
                    total_cells = table_df.shape[0] * table_df.shape[1]
                    non_empty_score = non_empty_cells / total_cells if total_cells > 0 else 0
                    table_score = (row_score * 0.4) + (col_score * 0.3) + (non_empty_score * 0.3)
                    
                    if table_score > best_score:
                        best_score = table_score
                        best_table = table_df
                except Exception as e:
                    self.logger.debug(f"【HTML处理】解析表格异常：{str(e)}")
                    continue
            
            if best_table is None:
                self.logger.error(f"【HTML处理】文件 {file_name} 无法解析任何表格")
                return None
            
            # 处理最佳表格
            df = best_table
            
            # 寻找最佳表头
            best_header_row = 0
            best_header_score = 0
            
            # 扫描前50行寻找表头
            for i in range(min(50, len(df))):
                header_candidate = df.iloc[i]
                standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                score = self._calculate_header_score(standardized_headers)
                
                if score > best_header_score:
                    best_header_score = score
                    best_header_row = i
            
            # 使用最佳表头行
            if best_header_row < len(df):
                best_header = df.iloc[best_header_row].apply(standardize_header)
                df.columns = best_header
                df = df.iloc[best_header_row + 1:].reset_index(drop=True)
                
                self.logger.info(f"【HTML表头识别完成】文件 {file_name} 最佳表头行：第{best_header_row+1}行")
                
                # 分块处理数据
                chunks = []
                total_rows = 0
                
                # 分块处理
                for i in range(0, len(df), self.chunk_size):
                    end = min(i + self.chunk_size, len(df))
                    chunk = df.iloc[i:end]
                    self.logger.info(f"【大文件处理】{file_name} 处理第{i//self.chunk_size+1}块，共{len(chunk)}行")
                    
                    processed_chunk = self._process_file_data(chunk, f"{file_name} - 块{i//self.chunk_size+1}")
                    if processed_chunk is not None and not processed_chunk.empty:
                        chunks.append(processed_chunk)
                        total_rows += len(processed_chunk)
                
                if chunks:
                    combined_df = pd.concat(chunks, ignore_index=True)
                    self.logger.info(f"【大文件处理】{file_name} 处理完成，共{total_rows}行数据")
                    return combined_df
                else:
                    self.logger.warning(f"【大文件处理】{file_name} 没有有效数据")
                    return None
            else:
                self.logger.warning(f"【HTML处理】文件 {file_name} 无法找到合适的表头")
                return None
        except Exception as e:
            self.logger.error(f"【大文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _preprocess_excel_file(self, file_path: str) -> str:
        """
        预处理Excel文件，取消冻结、隐藏、筛选，并恢复原始状态
        
        Args:
            file_path: 原始Excel文件路径
            
        Returns:
            预处理后的临时Excel文件路径
        """

        file_name = Path(file_path).name
        self.logger.debug(f"【Excel预处理】开始处理：{file_name}")
        
        try:
            # 导入文件类型检测函数
            from src.utils.file_utils import detect_file_type, get_file_extension_from_mime
            
            # 使用python-magic检测文件类型
            file_type = detect_file_type(file_path)
            file_ext = get_file_extension_from_mime(file_type)
            
            # 尝试使用openpyxl处理xlsx和xlsm文件
            if file_ext in ['.xlsx', '.xlsm']:
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter
                
                # 加载工作簿
                wb = load_workbook(file_path)
                
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    # 1. 取消冻结窗格
                    if sheet.freeze_panes != "A1":
                        sheet.freeze_panes = "A1"
                        self.logger.debug(f"【Excel预处理】{file_name} - {sheet_name}：取消冻结窗格")
                    
                    # 2. 取消筛选
                    if sheet.auto_filter is not None and sheet.auto_filter.ref:
                        sheet.auto_filter.ref = None
                        self.logger.debug(f"【Excel预处理】{file_name} - {sheet_name}：取消筛选")
                    
                    # 3. 显示所有隐藏的行
                    for row in sheet.iter_rows():
                        sheet.row_dimensions[row[0].row].hidden = False
                    
                    # 4. 显示所有隐藏的列
                    for col_idx in range(1, sheet.max_column + 1):
                        col_letter = get_column_letter(col_idx)
                        sheet.column_dimensions[col_letter].hidden = False
                    
                    # 5. 改进合并单元格的识别
                    # 记录所有合并单元格
                    merged_cells = list(sheet.merged_cells)
                    if merged_cells:
                        self.logger.debug(f"【Excel预处理】{file_name} - {sheet_name}：检测到 {len(merged_cells)} 个合并单元格")
                
                # 保存到临时文件
                import tempfile
                temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx', prefix='temp_bank_')
                os.close(temp_fd)
                try:
                    wb.save(temp_path)
                    return temp_path
                except Exception as e:
                    # 如果保存失败，清理临时文件
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    raise
            
            # 尝试使用xlrd处理xls文件
            elif file_ext == '.xls':
                # xlrd不支持修改文件，直接返回原文件
                self.logger.debug(f"【Excel预处理】{file_name}：XLS文件不支持修改，直接使用原文件")
                return file_path
            
            # 其他Excel格式
            else:
                self.logger.debug(f"【Excel预处理】{file_name}：不支持的Excel格式，直接使用原文件")
                return file_path
                
        except Exception as e:
            self.logger.error(f"【Excel预处理异常】{file_name}：{str(e)}")
            return file_path
    
    def process_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        处理单个文件，根据文件类型调用相应的处理方法
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        
        try:
            # 检查文件是否存在
            if not os.path.exists(file_path):
                self.logger.error(f"【文件不存在】{file_name}")
                return None
            
            # 获取文件缓存键（包含修改时间）
            cache_key = self._get_file_cache_key(file_path)
            
            # 检查文件结果缓存
            if cache_key in self.file_result_cache:

                # 更新访问时间
                self._update_cache_access("file", cache_key)
                return self.file_result_cache[cache_key]
            
            # 检查是否为大文件
            if self._is_large_file(file_path):
                self.logger.info(f"【文件处理】{file_name} 是大文件，使用分块处理")
                result = self._process_large_file(file_path)
            else:
                # 导入文件类型检测函数
                from src.utils.file_utils import is_text_file, is_pdf_file, is_excel_file
                
                # 处理TXT/CSV/HTML/HTM文件
                if is_text_file(file_path):
                    self.logger.debug(f"【文本文件处理】开始处理：{file_name}")
                    result = self._process_text_file(file_path)

                # 处理Excel文件（支持多sheet）
                elif is_excel_file(file_path):
                    self.logger.debug(f"【Excel文件处理】开始处理：{file_name}")
                    # 预处理Excel文件
                    processed_file_path = self._preprocess_excel_file(file_path)
                    try:
                        result = self._process_excel_file(processed_file_path)
                    finally:
                        # 清理临时文件
                        if processed_file_path != file_path:
                            if os.path.exists(processed_file_path):
                                try:
                                    os.remove(processed_file_path)
                                except Exception as e:
                                    self.logger.debug(f"【临时文件清理】无法删除文件 {processed_file_path}：{str(e)}")
                # 处理PDF文件
                elif is_pdf_file(file_path):
                    self.logger.debug(f"【PDF文件处理】开始处理：{file_name}")
                    result = self._process_pdf_file(file_path)
                else:
                    self.logger.error(f"【文件类型不支持】{file_name}：不支持的文件格式")
                    result = None
            
            # 缓存处理结果
            if result is not None:
                self.file_result_cache[cache_key] = result
                # 更新访问时间
                self._update_cache_access("file", cache_key)
                # 检查缓存大小
                self._cleanup_cache()
            
            return result
        except Exception as e:
            self.logger.error(f"【文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _is_abc_bank_html(self, soup) -> bool:
        """
        检测是否为农业银行HTML文件
        
        Args:
            soup: BeautifulSoup对象
            
        Returns:
            是否为农业银行HTML文件
        """
        # 检查标题是否包含农业银行标识
        title = ''
        if soup.title and soup.title.string:
            title = soup.title.string
        if '农业银行' in title:
            return True
        
        # 检查页面内容是否包含农业银行标识
        body_text = soup.get_text()
        if '中国农业银行' in body_text:
            return True
        
        # 检查表格结构
        tables = soup.find_all('table')
        if len(tables) >= 2:
            # 检查是否包含卡号、户名等基本信息
            for table in tables:
                table_text = table.get_text()
                if '卡号' in table_text or '户名' in table_text or '客户账号' in table_text:
                    return True
        
        return False
    
    def _process_abc_bank_html(self, file_path: str, encoding: str) -> Optional[pd.DataFrame]:
        """
        处理农业银行HTML文件
        
        Args:
            file_path: 文件路径
            encoding: 文件编码
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        
        try:
            from bs4 import BeautifulSoup
            import lxml
            
            # 使用BeautifulSoup结合lxml解析HTML
            with open(file_path, 'r', encoding=encoding) as f:
                soup = BeautifulSoup(f, 'lxml')
            
            # 检查是否为账户查询页面
            title = soup.title.string if soup.title else ''
            body_text = soup.get_text()
            
            if '客户账号查询' in title or '账户查询' in body_text:
                # 处理账户查询页面
                return self._process_abc_account_query_html(soup, file_name)
            else:
                # 处理交易流水页面
                # 提取账户信息
                account_info = self._extract_abc_account_info(soup)
                
                # 提取交易流水
                transaction_data = self._extract_abc_transactions(soup)
                
                if not transaction_data:
                    self.logger.warning(f"【农业银行HTML处理】文件 {file_name} 没有找到交易流水数据")
                    return None
                
                # 合并账户信息和交易流水
                all_transactions = []
                for account_no, transactions in transaction_data.items():
                    for transaction in transactions:
                        # 直接使用统一表头字段名
                        transaction['本方账号'] = account_no
                        # 添加账户信息
                        if account_no in account_info:
                            account = account_info[account_no]
                            transaction['本方账户开户行'] = account.get('开户行名称', '')
                            transaction['币种'] = account.get('币种', '')
                            transaction['户主名称'] = account.get('户主名称', '')
                            # 从产品类型中提取银行卡类别
                            product_type = account.get('产品类型', '')
                            if '借记卡' in product_type:
                                transaction['银行卡类别'] = '借记卡'
                            elif '信用卡' in product_type:
                                transaction['银行卡类别'] = '信用卡'
                            elif '一本通' in product_type:
                                transaction['银行卡类别'] = '一本通'
                            else:
                                transaction['银行卡类别'] = ''
                        # 优化对方账号/摘要字段的映射逻辑
                        counterparty_account = transaction.get('对方账号/摘要', '')
                        # 去除空格和符号，检查是否为纯数字
                        cleaned_account = ''.join(filter(str.isdigit, str(counterparty_account)))
                        if cleaned_account == str(counterparty_account).replace(' ', ''):
                            # 纯数字数据，映射到交易对方账号
                            transaction['交易对方账号'] = counterparty_account
                            # 如果摘要为空，使用原来的摘要字段
                            transaction['交易摘要'] = transaction.get('摘要', '')
                        else:
                            # 非纯数字数据，映射到交易摘要
                            transaction['交易对方账号'] = ''
                            # 合并摘要字段
                            if transaction.get('摘要', ''):
                                transaction['交易摘要'] = f"{counterparty_account} - {transaction.get('摘要', '')}"
                            else:
                                transaction['交易摘要'] = counterparty_account
                        transaction['交易对方名称'] = transaction.get('对方户名', '')
                        transaction['交易流水号'] = transaction.get('流水连续', '')
                        # 添加更多字段映射
                        transaction['交易地点'] = transaction.get('交易行名', '')
                        transaction['交易对方开户行'] = transaction.get('交易行名', '')
                        transaction['交易备注'] = transaction.get('摘要', '')
                    all_transactions.extend(transactions)
                
                # 转换为DataFrame
                df = pd.DataFrame(all_transactions)
                
                # 标准化表头
                df.columns = [standardize_header(col) for col in df.columns]
                
                # 处理交易金额列，确保为数值类型
                if '交易金额' in df.columns:
                    df['交易金额'] = pd.to_numeric(df['交易金额'], errors='coerce')
                
                # 处理余额列，确保为数值类型
                if '账户余额' in df.columns:
                    df['账户余额'] = pd.to_numeric(df['账户余额'], errors='coerce')
                
                # 添加借贷标志
                if '交易金额' in df.columns:
                    df['借贷标志'] = df['交易金额'].apply(lambda x: '收入' if x > 0 else '支出' if x < 0 else '')
                
                # 输出最终结果摘要
                self.logger.info(f"【农业银行HTML处理】成功处理 {file_name}")
                self.logger.info(f"【农业银行HTML处理】提取到 {len(account_info)} 个账户信息")
                if account_info:
                    # 收集所有唯一的户名
                    unique_names = set()
                    for acc_info in account_info.values():
                        name = acc_info.get('户主名称', '未知')
                        if name and name != '未知':
                            unique_names.add(name)
                    self.logger.info(f"【农业银行HTML处理】提取到的户名：{list(unique_names)}")
                self.logger.info(f"【农业银行HTML处理】提取到 {len(transaction_data)} 个账户的交易记录")
                total_transactions = sum(len(txs) for txs in transaction_data.values())
                self.logger.info(f"【农业银行HTML处理】总交易笔数：{total_transactions}")
                self.logger.info(f"【农业银行HTML处理】成功解析交易明细，共{len(df)}行数据")
                
                return df
        except Exception as e:
            self.logger.error(f"【农业银行HTML处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_abc_account_query_html(self, soup, file_name) -> Optional[pd.DataFrame]:
        """
        处理农业银行账户查询HTML页面
        
        Args:
            soup: BeautifulSoup对象
            file_name: 文件名
            
        Returns:
            处理后的DataFrame或None
        """
        try:
            # 提取户主名称
            户主名称 = ''
            tables = soup.find_all('table')
            
            # 常见姓氏和百家姓
            common_surnames = set([
                '赵', '钱', '孙', '李', '周', '吴', '郑', '王', '冯', '陈', '褚', '卫', '蒋', '沈', '韩', '杨',
                '朱', '秦', '尤', '许', '何', '吕', '施', '张', '孔', '曹', '严', '华', '金', '魏', '陶', '姜',
                '戚', '谢', '邹', '喻', '柏', '水', '窦', '章', '云', '苏', '潘', '葛', '奚', '范', '彭', '郎',
                '鲁', '韦', '昌', '马', '苗', '凤', '花', '方', '俞', '任', '袁', '柳', '酆', '鲍', '史', '唐',
                '费', '廉', '岑', '薛', '雷', '贺', '倪', '汤', '滕', '殷', '罗', '毕', '郝', '邬', '安', '常',
                '乐', '于', '时', '傅', '皮', '卞', '齐', '康', '伍', '余', '元', '卜', '顾', '孟', '平', '黄',
                '和', '穆', '萧', '尹', '姚', '邵', '湛', '汪', '祁', '毛', '禹', '狄', '米', '贝', '明', '臧',
                '计', '伏', '成', '戴', '谈', '宋', '茅', '庞', '熊', '纪', '舒', '屈', '项', '祝', '董', '梁',
                '杜', '阮', '蓝', '闵', '席', '季', '麻', '强', '贾', '路', '娄', '危', '江', '童', '颜', '郭',
                '梅', '盛', '林', '刁', '钟', '徐', '邱', '骆', '高', '夏', '蔡', '田', '樊', '胡', '凌', '霍',
                '虞', '万', '支', '柯', '昝', '管', '卢', '莫', '经', '房', '裘', '缪', '干', '解', '应', '宗',
                '丁', '宣', '贲', '邓', '郁', '单', '杭', '洪', '包', '诸', '左', '石', '崔', '吉', '钮', '龚',
                '程', '嵇', '邢', '滑', '裴', '陆', '荣', '翁', '荀', '羊', '于', '惠', '甄', '曲', '家', '封',
                '芮', '羿', '储', '靳', '汲', '邴', '糜', '松', '井', '段', '富', '巫', '乌', '焦', '巴', '弓',
                '牧', '隗', '山', '谷', '车', '侯', '宓', '蓬', '全', '郗', '班', '仰', '秋', '仲', '伊', '宫',
                '宁', '仇', '栾', '暴', '甘', '钭', '厉', '戎', '祖', '武', '符', '刘', '景', '詹', '束', '龙',
                '叶', '幸', '司', '韶', '郜', '黎', '蓟', '薄', '印', '宿', '白', '怀', '蒲', '邰', '从', '鄂',
                '索', '咸', '籍', '赖', '卓', '蔺', '屠', '蒙', '池', '乔', '阴', '郁', '胥', '能', '苍', '双',
                '闻', '莘', '党', '翟', '谭', '贡', '劳', '逄', '姬', '申', '扶', '堵', '冉', '宰', '郦', '雍',
                '却', '璩', '桑', '桂', '濮', '牛', '寿', '通', '边', '扈', '燕', '冀', '郏', '浦', '尚', '农',
                '温', '别', '庄', '晏', '柴', '瞿', '阎', '充', '慕', '连', '茹', '习', '宦', '艾', '鱼', '容',
                '向', '古', '易', '慎', '戈', '廖', '庾', '终', '暨', '居', '衡', '步', '都', '耿', '满', '弘',
                '匡', '国', '文', '寇', '广', '禄', '阙', '东', '欧', '殳', '沃', '利', '蔚', '越', '夔', '隆',
                '师', '巩', '厍', '聂', '晁', '勾', '敖', '融', '冷', '訾', '辛', '阚', '那', '简', '饶', '空',
                '曾', '毋', '沙', '乜', '养', '鞠', '须', '丰', '巢', '关', '蒯', '相', '查', '后', '荆', '红',
                '游', '竺', '权', '逯', '盖', '益', '桓', '公', '万俟', '司马', '上官', '欧阳', '夏侯', '诸葛',
                '闻人', '东方', '赫连', '皇甫', '尉迟', '公羊', '澹台', '申屠', '太叔', '轩辕', '令狐', '钟离',
                '宇文', '长孙', '慕容', '司寇', '仲孙'
            ])
            
            # 提取户主名称
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    for cell in cells:
                        cell_text = cell.get_text().strip()
                        if '户名' in cell_text:
                            # 查找户名值
                            next_cells = cell.find_next_siblings('td')
                            for next_cell in next_cells:
                                candidate_name = next_cell.get_text().strip()
                                if candidate_name and 2 <= len(candidate_name) <= 4:
                                    first_char = candidate_name[0]
                                    if first_char in common_surnames:
                                        户主名称 = candidate_name
                                        break
                            if 户主名称:
                                break
                    if 户主名称:
                        break
                if 户主名称:
                    break
            
            # 提取账户信息
            accounts = []
            current_branch = ''
            
            for table in tables:
                # 检查是否为开户行信息表格
                branch_rows = table.find_all('tr')
                if branch_rows and len(branch_rows) >= 2:
                    header_cells = branch_rows[0].find_all('td')
                    if header_cells and '开户行编号' in header_cells[0].get_text():
                        # 提取开户行名称
                        if len(branch_rows) > 1:
                            branch_cells = branch_rows[1].find_all('td')
                            if len(branch_cells) >= 2:
                                current_branch = branch_cells[1].get_text().strip()
                
                # 检查是否为账户信息表格
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    if len(cells) >= 8:
                        # 提取账户信息
                        product_type = cells[1].get_text().strip()
                        account_no = cells[2].get_text().strip()
                        # 提取账号，去除括号内的内容
                        account_no = re.sub(r'\(.*?\)', '', account_no).strip()
                        balance = cells[3].get_text().strip()
                        currency = cells[4].get_text().strip()
                        status = cells[5].get_text().strip()
                        open_date = cells[6].get_text().strip()
                        last_transaction_date = cells[7].get_text().strip()
                        
                        # 只处理有效的账号
                        if account_no and account_no != '/':
                            account = {
                                '户主名称': 户主名称,
                                '本方账号': account_no,
                                '本方账户开户行': current_branch,
                                '币种': currency,
                                '账户余额': balance,
                                '产品类型': product_type,
                                '状态': status,
                                '开户日期': open_date,
                                '最后交易日期': last_transaction_date
                            }
                            # 从产品类型中提取银行卡类别
                            if '借记卡' in product_type:
                                account['银行卡类别'] = '借记卡'
                            elif '信用卡' in product_type:
                                account['银行卡类别'] = '信用卡'
                            elif '一本通' in product_type:
                                account['银行卡类别'] = '一本通'
                            else:
                                account['银行卡类别'] = ''
                            accounts.append(account)
            
            if not accounts:
                self.logger.warning(f"【农业银行账户查询处理】文件 {file_name} 没有找到账户信息")
                return None
            
            # 转换为DataFrame
            df = pd.DataFrame(accounts)
            
            # 标准化表头
            df.columns = [standardize_header(col) for col in df.columns]
            
            # 处理账户余额列，确保为数值类型
            if '账户余额' in df.columns:
                df['账户余额'] = pd.to_numeric(df['账户余额'], errors='coerce')
            
            # 输出最终结果摘要
            self.logger.info(f"【农业银行账户查询处理】成功处理 {file_name}")
            self.logger.info(f"【农业银行账户查询处理】提取到 {len(accounts)} 个账户信息")
            self.logger.info(f"【农业银行账户查询处理】户主名称：{户主名称}")
            self.logger.info(f"【农业银行账户查询处理】成功解析账户信息，共{len(df)}行数据")
            
            return df
        except Exception as e:
            self.logger.error(f"【农业银行账户查询处理异常】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _extract_abc_account_info(self, soup) -> Dict[str, Dict[str, str]]:
        """
        提取农业银行HTML文件中的账户信息
        
        Args:
            soup: BeautifulSoup对象
            
        Returns:
            账户信息字典，键为客户账号，值为账户详细信息
        """
        account_info = {}
        户主名称 = ''
        
        # 查找所有包含账户信息的表格
        tables = soup.find_all('table')
        
        current_branch = ''
        
        # 常见姓氏和百家姓
        common_surnames = set([
            '赵', '钱', '孙', '李', '周', '吴', '郑', '王', '冯', '陈', '褚', '卫', '蒋', '沈', '韩', '杨',
            '朱', '秦', '尤', '许', '何', '吕', '施', '张', '孔', '曹', '严', '华', '金', '魏', '陶', '姜',
            '戚', '谢', '邹', '喻', '柏', '水', '窦', '章', '云', '苏', '潘', '葛', '奚', '范', '彭', '郎',
            '鲁', '韦', '昌', '马', '苗', '凤', '花', '方', '俞', '任', '袁', '柳', '酆', '鲍', '史', '唐',
            '费', '廉', '岑', '薛', '雷', '贺', '倪', '汤', '滕', '殷', '罗', '毕', '郝', '邬', '安', '常',
            '乐', '于', '时', '傅', '皮', '卞', '齐', '康', '伍', '余', '元', '卜', '顾', '孟', '平', '黄',
            '和', '穆', '萧', '尹', '姚', '邵', '湛', '汪', '祁', '毛', '禹', '狄', '米', '贝', '明', '臧',
            '计', '伏', '成', '戴', '谈', '宋', '茅', '庞', '熊', '纪', '舒', '屈', '项', '祝', '董', '梁',
            '杜', '阮', '蓝', '闵', '席', '季', '麻', '强', '贾', '路', '娄', '危', '江', '童', '颜', '郭',
            '梅', '盛', '林', '刁', '钟', '徐', '邱', '骆', '高', '夏', '蔡', '田', '樊', '胡', '凌', '霍',
            '虞', '万', '支', '柯', '昝', '管', '卢', '莫', '经', '房', '裘', '缪', '干', '解', '应', '宗',
            '丁', '宣', '贲', '邓', '郁', '单', '杭', '洪', '包', '诸', '左', '石', '崔', '吉', '钮', '龚',
            '程', '嵇', '邢', '滑', '裴', '陆', '荣', '翁', '荀', '羊', '于', '惠', '甄', '曲', '家', '封',
            '芮', '羿', '储', '靳', '汲', '邴', '糜', '松', '井', '段', '富', '巫', '乌', '焦', '巴', '弓',
            '牧', '隗', '山', '谷', '车', '侯', '宓', '蓬', '全', '郗', '班', '仰', '秋', '仲', '伊', '宫',
            '宁', '仇', '栾', '暴', '甘', '钭', '厉', '戎', '祖', '武', '符', '刘', '景', '詹', '束', '龙',
            '叶', '幸', '司', '韶', '郜', '黎', '蓟', '薄', '印', '宿', '白', '怀', '蒲', '邰', '从', '鄂',
            '索', '咸', '籍', '赖', '卓', '蔺', '屠', '蒙', '池', '乔', '阴', '郁', '胥', '能', '苍', '双',
            '闻', '莘', '党', '翟', '谭', '贡', '劳', '逄', '姬', '申', '扶', '堵', '冉', '宰', '郦', '雍',
            '却', '璩', '桑', '桂', '濮', '牛', '寿', '通', '边', '扈', '燕', '冀', '郏', '浦', '尚', '农',
            '温', '别', '庄', '晏', '柴', '瞿', '阎', '充', '慕', '连', '茹', '习', '宦', '艾', '鱼', '容',
            '向', '古', '易', '慎', '戈', '廖', '庾', '终', '暨', '居', '衡', '步', '都', '耿', '满', '弘',
            '匡', '国', '文', '寇', '广', '禄', '阙', '东', '欧', '殳', '沃', '利', '蔚', '越', '夔', '隆',
            '师', '巩', '厍', '聂', '晁', '勾', '敖', '融', '冷', '訾', '辛', '阚', '那', '简', '饶', '空',
            '曾', '毋', '沙', '乜', '养', '鞠', '须', '丰', '巢', '关', '蒯', '相', '查', '后', '荆', '红',
            '游', '竺', '权', '逯', '盖', '益', '桓', '公', '万俟', '司马', '上官', '欧阳', '夏侯', '诸葛',
            '闻人', '东方', '赫连', '皇甫', '尉迟', '公羊', '澹台', '申屠', '太叔', '轩辕', '令狐', '钟离',
            '宇文', '长孙', '慕容', '司寇', '仲孙'
        ])
        
        # 首先提取户主名称 - 使用新的提取方法
        # 查找全文件中第一个存在户名、户主名称、开户名称的行
        keyword_found = False
        for table in tables:
            rows = table.find_all('tr')
            for row_idx, row in enumerate(rows):
                cells = row.find_all('td')
                for cell_idx, cell in enumerate(cells):
                    cell_text = cell.get_text().strip()
                    # 精确匹配关键词
                    if cell_text in ['户名', '户主名称', '开户名称']:
                        keyword_found = True
                        candidates = []
                        
                        # 检查同列下一行 (n+1, n)
                        if row_idx + 1 < len(rows):
                            next_row = rows[row_idx + 1]
                            next_row_cells = next_row.find_all('td')
                            if cell_idx < len(next_row_cells):
                                candidate_name = next_row_cells[cell_idx].get_text().strip()
                                if candidate_name:
                                    candidates.append(candidate_name)
                        
                        # 检查同列下两行 (n+2, n)
                        if row_idx + 2 < len(rows):
                            next_next_row = rows[row_idx + 2]
                            next_next_row_cells = next_next_row.find_all('td')
                            if cell_idx < len(next_next_row_cells):
                                candidate_name = next_next_row_cells[cell_idx].get_text().strip()
                                if candidate_name:
                                    candidates.append(candidate_name)
                        
                        # 检查同行右侧第一列 (n, n+1)
                        if cell_idx + 1 < len(cells):
                            candidate_name = cells[cell_idx + 1].get_text().strip()
                            if candidate_name:
                                candidates.append(candidate_name)
                        
                        # 检查同行右侧第二列 (n, n+2)
                        if cell_idx + 2 < len(cells):
                            candidate_name = cells[cell_idx + 2].get_text().strip()
                            if candidate_name:
                                candidates.append(candidate_name)
                        
                        # 从候选中选择最合适的
                        if candidates:
                            # 优先选择人名
                            person_name = None
                            for candidate in candidates:
                                if 2 <= len(candidate) <= 4:
                                    first_char = candidate[0]
                                    if first_char in common_surnames:
                                        person_name = candidate
                                        break
                            
                            # 其次选择公司名
                            company_name = None
                            if not person_name:
                                for candidate in candidates:
                                    if 2 <= len(candidate) <= 20:
                                        if any(keyword in candidate for keyword in ['公司', '有限', '集团', '商行', '企业', '工作室']):
                                            company_name = candidate
                                            break
                            
                            # 确定最终的户主名称
                            if person_name:
                                户主名称 = person_name
                            elif company_name:
                                户主名称 = company_name
                            # 若都不合适，则留空
                            # 这里不设置户主名称，保持为空
                        
                        if 户主名称:
                            break
                if 户主名称:
                    break
            if 户主名称:
                break
        
        for table in tables:
            # 检查是否为开户行信息表格
            branch_rows = table.find_all('tr')
            if branch_rows and len(branch_rows) >= 2:
                header_cells = branch_rows[0].find_all('td')
                if header_cells and '开户行编号' in header_cells[0].get_text():
                    # 提取开户行名称
                    if len(branch_rows) > 1:
                        branch_cells = branch_rows[1].find_all('td')
                        if len(branch_cells) >= 2:
                            current_branch = branch_cells[1].get_text().strip()
            
            # 检查是否为账户信息表格
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 5:
                    # 提取客户账号、产品类型、账户余额、币种
                    product_type = cells[1].get_text().strip()
                    account_no = cells[2].get_text().strip()
                    # 提取账号，去除括号内的内容
                    account_no = re.sub(r'\(.*?\)', '', account_no).strip()
                    balance = cells[3].get_text().strip()
                    currency = cells[4].get_text().strip()
                    
                    # 只处理有效的账号
                    if account_no and account_no != '/':
                        account_info[account_no] = {
                            '产品类型': product_type,
                            '账户余额': balance,
                            '币种': currency,
                            '开户行名称': current_branch,
                            '户主名称': 户主名称
                        }
        
        return account_info
    
    def _extract_abc_transactions(self, soup) -> Dict[str, List[Dict[str, str]]]:
        """
        提取农业银行HTML文件中的交易流水
        
        Args:
            soup: BeautifulSoup对象
            
        Returns:
            交易流水字典，键为客户账号，值为交易记录列表
        """
        transactions = {}
        
        # 查找交易明细部分
        detail_div = soup.find('div', id='detailDiv')
        if not detail_div:
            return transactions
        
        # 查找所有表格
        all_tables = detail_div.find_all('table')
        
        current_account = ''
        current_transactions = []
        
        for table in all_tables:
            # 检查是否为客户账号标题表格
            header_rows = table.find_all('tr')
            if header_rows:
                header_cell = header_rows[0].find('td')
                if header_cell:
                    header_text = header_cell.get_text()
                    if '客户账号' in header_text:
                        # 提取客户账号
                        account_match = re.search(r'客户账号\s*(\d+)', header_text)
                        if account_match:
                            account_no = account_match.group(1)
                            # 保存之前的交易记录
                            if current_account and current_transactions:
                                transactions[current_account] = current_transactions
                            # 开始新的账号交易记录
                            current_account = account_no
                            # 如果该账号已经存在交易记录，追加而不是重置
                            if current_account in transactions:
                                current_transactions = transactions[current_account]
                            else:
                                current_transactions = []
                            continue
            
            # 检查是否为交易流水表格
            rows = table.find_all('tr')
            if len(rows) >= 2:
                # 检查表头
                header_row = rows[0]
                header_cells = header_row.find_all('td')
                if header_cells:
                    # 检查是否包含交易日期列
                    has_transaction_date = any('交易日期' in cell.get_text() for cell in header_cells)
                    if has_transaction_date:
                        # 提取交易记录
                        for row in rows[1:]:
                            cells = row.find_all('td')
                            if len(cells) >= 11:
                                # 提取交易日期
                                date_time_cell = cells[0].get_text().strip()
                                date_time_parts = date_time_cell.split('\n')
                                date_str = date_time_parts[0].strip() if date_time_parts else ''
                                
                                # 转换日期格式
                                transaction_date = self._format_abc_date(date_str)
                                
                                # 提取其他字段
                                transaction_amount = cells[1].get_text().strip()
                                account_balance = cells[2].get_text().strip()
                                transaction_branch_code = cells[3].get_text().strip()
                                transaction_branch_name = cells[4].get_text().strip()
                                voucher_no = cells[5].get_text().strip()
                                transaction_code = cells[6].get_text().strip()
                                counterparty_account = cells[7].get_text().strip()
                                counterparty_name = cells[8].get_text().strip()
                                summary = cells[9].get_text().strip()
                                serial_no = cells[10].get_text().strip()
                                
                                # 创建交易记录
                                transaction = {
                                    '交易日期': transaction_date,
                                    '交易金额': transaction_amount,
                                    '账户余额': account_balance,
                                    '交易行号': transaction_branch_code,
                                    '交易行名': transaction_branch_name,
                                    '传票号': voucher_no,
                                    '交易代码': transaction_code,
                                    '对方账号/摘要': counterparty_account,
                                    '对方户名': counterparty_name,
                                    '摘要': summary,
                                    '流水连续': serial_no
                                }
                                
                                current_transactions.append(transaction)
        
        # 保存最后一个账号的交易记录
        if current_account and current_transactions:
            transactions[current_account] = current_transactions
        
        return transactions
    
    def _format_abc_date(self, date_str: str) -> str:
        """
        格式化农业银行日期
        
        Args:
            date_str: 原始日期字符串，格式为YYYYMMDD或YYYYMMDDHHMM
            
        Returns:
            格式化后的日期字符串，格式为YYYY-MM-DD
        """
        if not date_str or date_str == '/':
            return ''
        
        try:
            # 提取纯日期部分（前8位）
            pure_date = date_str[:8]
            if len(pure_date) == 8:
                return f"{pure_date[:4]}-{pure_date[4:6]}-{pure_date[6:]}"
        except Exception:
            pass
        
        return date_str
    
    def _format_abc_time(self, time_str: str) -> str:
        """
        格式化农业银行时间
        
        Args:
            time_str: 原始时间字符串，格式为HHMM
            
        Returns:
            格式化后的时间字符串，格式为HH:MM:00
        """
        if not time_str or time_str == '/':
            return ''
        
        try:
            time_str = time_str.replace('&nbsp;', '').strip()
            if len(time_str) == 4:
                return f"{time_str[:2]}:{time_str[2:]}:00"
        except Exception:
            pass
        
        return time_str
    
    def _process_text_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        处理文本文件（TXT/CSV/HTML/HTM）
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        
        try:
            # 导入文件类型检测函数
            from src.utils.file_utils import detect_file_type, get_file_extension_from_mime, magic_available
            
            # 检测文件编码（使用缓存）
            if file_path not in self.encoding_cache:
                self.encoding_cache[file_path] = detect_file_encoding(file_path, self.config)
            encoding = self.encoding_cache[file_path]
            # 更新访问时间
            self._update_cache_access("encoding", file_path)
            self.logger.debug(f"【文本文件处理】{file_name} 检测到编码：{encoding}")
            
            # 使用python-magic检测文件类型
            file_type = detect_file_type(file_path)
            file_ext = get_file_extension_from_mime(file_type)
            
            # 判断是否为HTML文件
            is_html = False
            if magic_available:
                is_html = file_type == 'text/html'
            else:
                is_html = file_path.lower().endswith(('.html', '.htm'))
            
            if is_html:
                # 处理HTML文件
                self.logger.debug(f"【HTML文件处理】开始处理：{file_name}")
                try:
                    from bs4 import BeautifulSoup
                    import lxml
                    
                    # 使用BeautifulSoup结合lxml解析HTML
                    with open(file_path, 'r', encoding=encoding) as f:
                        soup = BeautifulSoup(f, 'lxml')
                    
                    # 检测是否为农业银行HTML文件
                    if self._is_abc_bank_html(soup):
                        # 特殊处理农业银行HTML文件
                        return self._process_abc_bank_html(file_path, encoding)
                    
                    # 提取所有表格
                    tables = soup.find_all('table')
                    
                    if not tables:
                        self.logger.error(f"【HTML处理】文件 {file_name} 没有找到表格")
                        return None
                    
                    # 表格评分和选择
                    best_table = None
                    best_score = 0
                    
                    for table in tables:
                        try:
                            # 转换为DataFrame，解决FutureWarning问题
                            from io import StringIO
                            table_html = str(table)
                            try:
                                # 使用更宽松的匹配模式，添加flavor参数以提高解析成功率
                                table_dfs = pd.read_html(
                                    StringIO(table_html), 
                                    header=None, 
                                    match='.*',  # 更宽松的匹配模式
                                    flavor=['lxml', 'html5lib']  # 尝试多种解析器
                                )
                                if not table_dfs:
                                    continue
                            except Exception as e:
                                # 当表格解析失败时，尝试创建一个空DataFrame作为备选
                                self.logger.debug(f"【HTML处理】表格解析失败，创建空DataFrame：{str(e)}")
                                continue
                            
                            table_df = table_dfs[0]
                            
                            # 计算表格得分
                            # 1. 行数得分（行数越多得分越高，但不过分偏好非常大的表格）
                            row_score = min(len(table_df), 100) / 100
                            
                            # 2. 列数得分（列数适中得分高，10-20列最佳）
                            col_count = len(table_df.columns)
                            if 10 <= col_count <= 20:
                                col_score = 1.0
                            elif 5 <= col_count < 10 or 20 < col_count <= 30:
                                col_score = 0.7
                            elif col_count < 5 or col_count > 30:
                                col_score = 0.3
                            else:
                                col_score = 0.5
                            
                            # 3. 非空单元格比例得分
                            non_empty_cells = table_df.notna().sum().sum()
                            total_cells = table_df.shape[0] * table_df.shape[1]
                            non_empty_score = non_empty_cells / total_cells if total_cells > 0 else 0
                            
                            # 综合得分
                            table_score = (row_score * 0.4) + (col_score * 0.3) + (non_empty_score * 0.3)
                            
                            # 如果当前表格更好，更新最佳表格
                            if table_score > best_score:
                                best_score = table_score
                                best_table = table_df
                        except Exception as e:
                            self.logger.debug(f"【HTML处理】解析表格异常：{str(e)}")
                            continue
                    
                    if best_table is None:
                        self.logger.error(f"【HTML处理】文件 {file_name} 无法解析任何表格")
                        return None
                    
                    # 在表格前50行和模拟的第2至5页前30行中寻找最佳表头
                    df = best_table
                    best_header_row = 0
                    best_header_score = 0
                    
                    # 定义要扫描的行范围
                    scan_ranges = [
                        (0, 50),      # 第1页：前50行
                        (50, 80),     # 第2页：51-80行（前30行）
                        (100, 130),   # 第3页：101-130行（前30行）
                        (150, 180),   # 第4页：151-180行（前30行）
                        (200, 230)    # 第5页：201-230行（前30行）
                    ]
                    
                    self.logger.debug(f"【HTML表头识别】开始扫描多个行范围寻找最佳表头")
                    
                    # 遍历所有扫描范围
                    for page_num, (start_row, end_row) in enumerate(scan_ranges, 1):
                        # 确保扫描范围不超出数据范围
                        actual_start = max(0, start_row)
                        actual_end = min(end_row, len(df))
                        
                        if actual_start >= actual_end:
                            continue
                        
                        self.logger.debug(f"【HTML表头识别】开始扫描第{page_num}页的第{actual_start+1}-{actual_end}行")
                        
                        for i in range(actual_start, actual_end):
                            # 尝试将第i行作为表头
                            header_candidate = df.iloc[i]
                            
                            # 标准化表头候选
                            standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                            
                            # 计算表头得分
                            total_score = self._calculate_header_score(standardized_headers)
                            exact_match_score = sum(1 for h in standardized_headers if h in self.header_mapping)
                            keyword_match_score = 0
                            for h in standardized_headers:
                                for unified_col, keywords in self.keyword_mapping.items():
                                    if any(keyword in h for keyword in keywords):
                                        keyword_match_score += 1
                                        break
                            
                            # 更新最佳表头：只在得分更高时更新，或者得分相同时保留更前面的行
                            if total_score > best_header_score:
                                best_header_score = total_score
                                best_header_row = i
                    
                    # 使用最佳表头行
                    if best_header_row < len(df):
                        # 获取最佳表头
                        best_header = df.iloc[best_header_row].apply(standardize_header)
                        df.columns = best_header
                        df = df.iloc[best_header_row + 1:].reset_index(drop=True)
                        
                        self.logger.info(f"【HTML表头识别完成】文件 {file_name} 最佳表头行：第{best_header_row+1}行")
                        self.logger.info(f"【HTML表头识别完成】最佳表头：{best_header.to_list()}")
                        self.logger.info(f"【HTML表头识别完成】表头得分：{best_header_score}")
                        print(f"【HTML表头识别完成】文件 {file_name} 最佳表头行：第{best_header_row+1}行")
                        print(f"【HTML表头识别完成】最佳表头：{best_header.to_list()}")
                    else:
                        # 如果没有找到合适的表头，返回None，放入待复核文件
                        self.logger.warning(f"【HTML处理】文件 {file_name} 无法找到合适的表头，放入待复核文件")
                        return None
                    
                except Exception as e:
                    self.logger.error(f"【HTML处理异常】{file_name}：{str(e)}", exc_info=True)
                    return None
            else:
                # 处理CSV/TXT文件
                # 判断是否为CSV文件
                is_csv = False
                if magic_available:
                    is_csv = file_type in ['text/csv', 'application/csv']
                else:
                    is_csv = file_path.lower().endswith('.csv')
                
                file_type_name = "CSV" if is_csv else "TXT"
                self.logger.debug(f"【{file_type_name}文件处理】开始处理：{file_name}")
                
                # 确定文件路径（TXT文件需要预处理）
                actual_file_path = file_path
                if not is_csv:
                    actual_file_path = preprocess_txt_file(file_path, encoding, self.config)
                
                # 定义常见分隔符列表
                common_delimiters = [',', '\t', ';', '|', ' ']
                
                # 检测分隔符（使用缓存）
                if actual_file_path not in self.delimiter_cache:
                    self.delimiter_cache[actual_file_path] = detect_file_delimiter(actual_file_path, encoding, self.config)
                detected_delimiter = self.delimiter_cache[actual_file_path]
                # 更新访问时间
                self._update_cache_access("delimiter", actual_file_path)
                
                self.logger.debug(f"【{file_type_name}文件处理】{file_name} 检测到分隔符：{repr(detected_delimiter)}")
                
                # 准备分隔符策略列表
                delimiter_strategies = []
                
                # 添加检测到的分隔符
                if detected_delimiter in common_delimiters:
                    delimiter_strategies.append(detected_delimiter)
                else:
                    # 如果检测到的分隔符不在常见列表中，也添加它
                    delimiter_strategies.append(detected_delimiter)
                
                # 添加所有常见分隔符作为备选策略
                for delim in common_delimiters:
                    if delim not in delimiter_strategies:
                        delimiter_strategies.append(delim)
                
                self.logger.debug(f"【{file_type_name}文件处理】{file_name} 准备尝试的分隔符策略：{[repr(d) for d in delimiter_strategies]}")
                
                # 尝试多种编码读取文件
                encodings = [encoding, 'gbk', 'utf-8', 'latin-1']
                
                # 首先尝试多账户处理
                # 先尝试使用检测到的分隔符读取原始数据
                df_raw = None
                for enc in encodings:
                    try:
                        # 首先使用检测到的分隔符
                        df_raw = pd.read_csv(
                            actual_file_path,
                            encoding=enc,
                            sep=detected_delimiter,
                            header=None,
                            low_memory=False,
                            skip_blank_lines=True,
                            on_bad_lines='skip'
                        )
                        if df_raw is not None and not df_raw.empty:
                            break
                    except Exception:
                        try:
                            # 如果失败，尝试自动检测分隔符
                            df_raw = pd.read_csv(
                                actual_file_path,
                                encoding=enc,
                                sep=None,  # 自动检测
                                header=None,
                                low_memory=False,
                                skip_blank_lines=True,
                                on_bad_lines='skip'
                            )
                            if df_raw is not None and not df_raw.empty:
                                break
                        except Exception:
                            continue
                
                if df_raw is not None and not df_raw.empty:
                    # 首先在数据前50行中寻找最佳表头
                    best_header_row = 0
                    best_header_score = 0
                    best_header_headers = []
                    
                    # 最多检查前50行
                    max_check_rows = min(50, len(df_raw))
                    
                    self.logger.debug(f"【文本文件表头识别】开始扫描前{max_check_rows}行寻找最佳表头")
                    
                    for i in range(max_check_rows):
                        # 尝试将第i行作为表头
                        header_candidate = df_raw.iloc[i]
                        
                        # 标准化表头候选
                        standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                        
                        # 计算表头得分
                        total_score = self._calculate_header_score(standardized_headers)
                        
                        # 更新最佳表头：只在得分更高时更新，得分相同时保留更前面的行
                        if total_score > best_header_score:
                            best_header_score = total_score
                            best_header_row = i
                            best_header_headers = standardized_headers
                    
                    # 输出最佳表头信息
                    self.logger.info(f"【表头分析】{file_name} 最佳表头行：第{best_header_row+1}行，得分：{best_header_score}")
                    self.logger.info(f"【表头分析】最佳表头：{best_header_headers}")
                    
                    # 使用最佳表头行
                    if best_header_row < len(df_raw):
                        df = df_raw.copy()
                        # 获取最佳表头
                        best_header = df.iloc[best_header_row].apply(standardize_header)
                        df.columns = best_header
                        df = df.iloc[best_header_row + 1:].reset_index(drop=True)
                        
                        # 处理文件数据，进行表头映射（临时映射，不进行关键列检查）
                        self.logger.info(f"【表头映射】{file_name} 开始临时表头映射")
                        
                        # 创建临时映射函数，跳过关键列检查
                        def temp_map(df_temp, file_id):
                            # 表头标准化
                            original_raw_headers = list(df_temp.columns)
                            df_temp.columns = [standardize_header(col) for col in df_temp.columns]
                            standardized_headers = list(df_temp.columns)
                            
                            # 优化重复列
                            df_temp = optimize_duplicate_columns(df_temp)
                            standardized_headers = list(df_temp.columns)
                            
                            # 映射到统一表头
                            mapped_df = pd.DataFrame(columns=self.unified_headers)
                            matched_cols = []
                            
                            # 收集所有可能的匹配
                            all_matches = []
                            
                            # 1. 收集所有精确匹配
                            for col in standardized_headers:
                                if col in self.header_mapping:
                                    unified_col = self.header_mapping[col]
                                    mapping_items = list(self.header_mapping.items())
                                    position = next((i for i, (k, v) in enumerate(mapping_items) if k == col), len(mapping_items))
                                    position_weight = max(0, 10 - position * 0.1)
                                    weight = 20 + position_weight
                                    
                                    all_matches.append({
                                        'original': col,
                                        'matched': unified_col,
                                        'method': '精确匹配',
                                        'weight': weight,
                                        'position': position
                                    })
                            
                            # 2. 收集所有关键词匹配
                            for col in standardized_headers:
                                if any(m['original'] == col and m['method'] == '精确匹配' for m in all_matches):
                                    continue
                                
                                for unified_col, keywords in self.keyword_mapping.items():
                                    for i, keyword in enumerate(keywords):
                                        if unified_col in self.exclusion_rules:
                                            exclusion_keywords = self.exclusion_rules[unified_col]
                                            if any(exclusion_keyword in col for exclusion_keyword in exclusion_keywords):
                                                continue
                                        
                                        score = 0
                                        if col == keyword:
                                            score = 10
                                        elif col.startswith(keyword) or col.endswith(keyword):
                                            score = 8
                                        elif keyword in col:
                                            keyword_chars = set(keyword)
                                            col_chars = set(col)
                                            if keyword_chars:
                                                common_chars = keyword_chars.intersection(col_chars)
                                                coverage_ratio = len(common_chars) / len(keyword_chars)
                                                if coverage_ratio > 0.5:
                                                    score = 5
                                        
                                        if score > 0:
                                            keyword_position_weight = max(0, 5 - i * 0.5)
                                            weight = score + keyword_position_weight
                                            
                                            all_matches.append({
                                                'original': col,
                                                'matched': unified_col,
                                                'method': '关键词匹配',
                                                'keyword': keyword,
                                                'weight': weight,
                                                'score': score,
                                                'keyword_position': i
                                            })
                            
                            # 3. 处理匹配冲突
                            matches_by_original = {}
                            for match in all_matches:
                                original_col = match['original']
                                if original_col not in matches_by_original:
                                    matches_by_original[original_col] = []
                                matches_by_original[original_col].append(match)
                            
                            selected_matches = []
                            for original_col, matches in matches_by_original.items():
                                sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                                best_match = sorted_matches[0]
                                selected_matches.append(best_match)
                            
                            # 4. 再次处理冲突
                            matches_by_unified = {}
                            for match in selected_matches:
                                unified_col = match['matched']
                                if unified_col not in matches_by_unified:
                                    matches_by_unified[unified_col] = []
                                matches_by_unified[unified_col].append(match)
                            
                            final_selected_matches = []
                            for unified_col, matches in matches_by_unified.items():
                                sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                                best_match = sorted_matches[0]
                                final_selected_matches.append(best_match)
                            
                            selected_matches = final_selected_matches
                            
                            # 应用选定的匹配
                            for match in selected_matches:
                                original_col = match['original']
                                unified_col = match['matched']
                                if unified_col in mapped_df.columns:
                                    mapped_df[unified_col] = df_temp[original_col]
                                    matched_cols.append(match)
                            
                            # 排除全NA列
                            mapped_df = mapped_df.dropna(axis=1, how='all')
                            
                            # 将NaN替换为空字符串
                            mapped_df = mapped_df.fillna('').infer_objects()
                            
                            return mapped_df if not mapped_df.empty else None
                        
                        # 执行临时映射
                        temp_result = temp_map(df, f"{file_name} - 临时映射")
                        self.logger.info(f"【表头映射】{file_name} 临时表头映射完成，结果：{'成功' if temp_result is not None else '失败'}")
                        
                        # 检查表头映射结果
                        if temp_result is not None:
                            # 检查缺失的统一表头列
                            mapped_headers = list(temp_result.columns)
                            missing_headers = [col for col in self.unified_headers if col not in mapped_headers]
                            
                            self.logger.info(f"【表头映射分析】{file_name} 成功映射的统一表头列：{mapped_headers}")
                            self.logger.info(f"【表头映射分析】{file_name} 缺失的统一表头列：{missing_headers}")
                            
                            # 检查户主名称列是否在缺失列表中
                            has_missing_account_name = '户主名称' in missing_headers
                            self.logger.info(f"【表头映射分析】{file_name} 户主名称列是否缺失：{has_missing_account_name}")
                            
                            # 检查是否存在其他表头行评分大于等于最佳表头行评分
                            has_other_header_rows = False
                            if has_missing_account_name:
                                # 寻找最佳表头
                                current_best_header_row = 0
                                current_best_header_score = 0
                                
                                # 读取前50行寻找表头
                                header_df = pd.read_csv(
                                    actual_file_path,
                                    encoding=encoding,
                                    sep=delimiter,
                                    header=None,
                                    nrows=50,
                                    skip_blank_lines=True,
                                    on_bad_lines='skip'
                                )
                                
                                for i in range(len(header_df)):
                                    header_candidate = header_df.iloc[i]
                                    standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                                    score = self._calculate_header_score(standardized_headers)
                                    
                                    if score > current_best_header_score:
                                        current_best_header_score = score
                                        current_best_header_row = i
                                
                                # 检查是否存在其他表头行评分大于等于最佳表头行评分
                                for i in range(len(header_df)):
                                    if i == current_best_header_row:
                                        continue
                                    header_candidate = header_df.iloc[i]
                                    standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                                    score = self._calculate_header_score(standardized_headers)
                                    if score >= current_best_header_score:
                                        has_other_header_rows = True
                                        break
                            
                            # 如果缺失的统一表头列中包含户主名称且存在其他表头行，尝试多账户处理
                            if has_missing_account_name and has_other_header_rows:
                                self.logger.info(f"【表头分析】{file_name} 表头映射后缺失户主名称列且存在其他表头行，尝试多账户处理")
                                self.logger.info(f"【多账户处理】开始多账户处理流程")
                                multi_account_df = self.multi_account_handler.process_multi_account_sheet(
                                    df_raw, 
                                    file_name
                                )
                                
                                if multi_account_df is not None:
                                    # 多账户处理成功，直接处理结果
                                    self.logger.info(f"【多账户处理成功】{file_name} 采用多账户处理方式")
                                    return self._process_file_data(multi_account_df, file_name)
                                else:
                                    self.logger.info(f"【多账户处理】多账户处理失败，继续使用原处理流程")
                            else:
                                if has_missing_account_name:
                                    self.logger.info(f"【表头分析】{file_name} 表头映射后缺失户主名称列，但不存在其他表头行，使用原处理流程")
                                else:
                                    self.logger.info(f"【表头分析】{file_name} 表头映射后不缺失户主名称列，使用原处理流程")
                        else:
                            self.logger.warning(f"【表头映射】{file_name} 表头映射失败，继续使用原处理流程")
                
                # 对每个分隔符策略进行评估
                best_strategy = None
                best_score = -1
                best_df = None
                best_header_row = -1
                best_header = None
                
                for delimiter in delimiter_strategies:
                    self.logger.debug(f"【{file_type_name}文件处理】{file_name} 尝试分隔符策略：{repr(delimiter)}")
                    
                    # 读取文件并处理编码问题
                    def read_with_encoding(enc, skip_rows=0):
                        try:
                            return pd.read_csv(
                                actual_file_path,
                                encoding=enc,
                                sep=delimiter,
                                header=None,
                                low_memory=False,
                                skip_blank_lines=True,
                                on_bad_lines='skip',  # 跳过坏行，提高容错性
                                skiprows=skip_rows
                            )
                        except Exception:
                            return None
                    
                    # 尝试读取文件
                    df_raw = None
                    for enc in encodings:
                        df_raw = read_with_encoding(enc)
                        if df_raw is not None and not df_raw.empty:
                            break
                    
                    # 检查读取结果是否异常
                    file_size = os.path.getsize(actual_file_path)
                    is_abnormal = False
                    
                    if df_raw is None or df_raw.empty:
                        is_abnormal = True
                    else:
                        # 检查行列数与文件大小是否匹配
                        # 小文件（<1KB）可以接受较少的行列数
                        if file_size > 1024:
                            if len(df_raw) < 3 or len(df_raw.columns) < 2:
                                is_abnormal = True
                            # 检查空值比例，超过80%视为异常
                            if df_raw.isnull().sum().sum() / (len(df_raw) * len(df_raw.columns)) > 0.8:
                                is_abnormal = True
                    
                    # 增强读取逻辑
                    if is_abnormal:
                        self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 常规读取异常，尝试增强读取")
                        
                        # 阶段1：跳过前N行（最多50行）
                        enhanced_df = None
                        for skip in range(1, 51):
                            for enc in encodings:
                                enhanced_df = read_with_encoding(enc, skip_rows=skip)
                                if enhanced_df is not None and not enhanced_df.empty:
                                    if len(enhanced_df.columns) >= 2 and enhanced_df.isnull().sum().sum() / (len(enhanced_df) * len(enhanced_df.columns)) <= 0.8:
                                        self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 跳过前{skip}行后读取成功")
                                        df_raw = enhanced_df
                                        break
                            if enhanced_df is not None and not enhanced_df.empty and len(enhanced_df.columns) >= 2:
                                break
                        
                        # 阶段2：跳过第一页（前50行），从第二页开始读取
                        if enhanced_df is None or enhanced_df.empty or len(enhanced_df.columns) < 2:
                            self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 跳过前50行后仍异常，尝试跳过第一页")
                            
                            # 第二页从第50行开始，尝试跳过前30行
                            for skip in range(50, 80):
                                for enc in encodings:
                                    enhanced_df = read_with_encoding(enc, skip_rows=skip)
                                    if enhanced_df is not None and not enhanced_df.empty:
                                        if len(enhanced_df.columns) >= 2 and enhanced_df.isnull().sum().sum() / (len(enhanced_df) * len(enhanced_df.columns)) <= 0.8:
                                            self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 跳过第一页后读取成功，起始行：{skip}")
                                            df_raw = enhanced_df
                                            break
                                if enhanced_df is not None and not enhanced_df.empty and len(enhanced_df.columns) >= 2:
                                    break
                        
                        # 阶段3：跳过第一页和第二页前30行，从第二页31行开始读取
                        if enhanced_df is None or enhanced_df.empty or len(enhanced_df.columns) < 2:
                            self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 跳过第二页前30行后仍异常，尝试跳过更多行")
                            
                            # 从第二页31行开始
                            for skip in range(80, 110):
                                for enc in encodings:
                                    enhanced_df = read_with_encoding(enc, skip_rows=skip)
                                    if enhanced_df is not None and not enhanced_df.empty:
                                        if len(enhanced_df.columns) >= 2 and enhanced_df.isnull().sum().sum() / (len(enhanced_df) * len(enhanced_df.columns)) <= 0.8:
                                            self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 跳过更多行后读取成功，起始行：{skip}")
                                            df_raw = enhanced_df
                                            break
                                if enhanced_df is not None and not enhanced_df.empty and len(enhanced_df.columns) >= 2:
                                    break
                    
                    # 最终检查
                    if df_raw is None or df_raw.empty:
                        self.logger.debug(f"【{file_type_name}文件处理】{file_name} 使用分隔符 {repr(delimiter)} 无法读取有效数据，跳过此策略")
                        continue
                    
                    # 在文件前50行和模拟的第2至5页前30行中寻找最佳表头
                    current_best_header_row = 0
                    current_best_score = 0
                    
                    # 定义要扫描的行范围
                    scan_ranges = [
                        (0, 50),      # 第1页：前50行
                        (50, 80),     # 第2页：51-80行（前30行）
                        (100, 130),   # 第3页：101-130行（前30行）
                        (150, 180),   # 第4页：151-180行（前30行）
                        (200, 230)    # 第5页：201-230行（前30行）
                    ]
                    
                    self.logger.debug(f"【表头识别】使用分隔符 {repr(delimiter)} 开始扫描多个行范围寻找最佳表头")
                    
                    # 遍历所有扫描范围
                    for page_num, (start_row, end_row) in enumerate(scan_ranges, 1):
                        # 确保扫描范围不超出数据范围
                        actual_start = max(0, start_row)
                        actual_end = min(end_row, len(df_raw))
                        
                        if actual_start >= actual_end:
                            continue
                        
                        self.logger.debug(f"【表头识别】使用分隔符 {repr(delimiter)} 开始扫描第{page_num}页的第{actual_start+1}-{actual_end}行")
                        
                        for i in range(actual_start, actual_end):
                            # 尝试将第i行作为表头
                            header_candidate = df_raw.iloc[i]
                            
                            # 标准化表头候选
                            standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                            
                            # 计算表头得分
                            total_score = self._calculate_header_score(standardized_headers)
                            exact_match_score = sum(1 for h in standardized_headers if h in self.header_mapping)
                            keyword_match_score = 0
                            for h in standardized_headers:
                                for unified_col, keywords in self.keyword_mapping.items():
                                    if any(keyword in h for keyword in keywords):
                                        keyword_match_score += 1
                                        break
                            
                            # 更新最佳表头：只在得分更高时更新，得分相同时保留更前面的行
                            if total_score > current_best_score:
                                current_best_score = total_score
                                current_best_header_row = i
                    
                    # 使用最佳表头行
                    if current_best_header_row < len(df_raw):
                        # 获取最佳表头
                        current_header = df_raw.iloc[current_best_header_row].apply(standardize_header)
                        current_df = df_raw.copy()
                        current_df.columns = current_header
                        current_df = current_df.iloc[current_best_header_row + 1:].reset_index(drop=True)
                        
                        self.logger.debug(f"【{file_type_name}表头识别】文件 {file_name} 使用分隔符 {repr(delimiter)} 最佳表头行：第{current_best_header_row+1}行，得分：{current_best_score}")
                        
                        # 比较得分，选择最佳策略
                        if current_best_score > best_score:
                            best_score = current_best_score
                            best_strategy = delimiter
                            best_df = current_df
                            best_header_row = current_best_header_row
                            best_header = current_header
                
                # 检查是否找到最佳策略
                if best_strategy is None or best_df is None:
                    self.logger.warning(f"【{file_type_name}文件处理】{file_name} 所有分隔符策略均无法读取有效数据")
                    return None
                
                # 使用最佳策略
                delimiter = best_strategy
                df = best_df
                
                self.logger.info(f"【{file_type_name}文件处理】{file_name} 选择最佳分隔符策略：{repr(delimiter)}，表头得分：{best_score}")
                self.logger.info(f"【{file_type_name}表头识别完成】文件 {file_name} 最佳表头行：第{best_header_row+1}行")
                self.logger.info(f"【{file_type_name}表头识别完成】最佳表头：{best_header.to_list()}")
                print(f"【{file_type_name}文件处理】文件 {file_name} 选择最佳分隔符策略：{repr(delimiter)}")
                print(f"【{file_type_name}表头识别完成】文件 {file_name} 最佳表头行：第{best_header_row+1}行")
                print(f"【{file_type_name}表头识别完成】最佳表头：{best_header.to_list()}")
                
                # 更新分隔符缓存为最佳策略
                self.delimiter_cache[actual_file_path] = best_strategy
                self._update_cache_access("delimiter", actual_file_path)
            
            self.logger.debug(f"【文本文件处理】{file_name} 读取成功，共 {len(df)} 行，{len(df.columns)} 列")
            
            # 添加内容验证：检查是否包含银行流水所需的关键信息
            df.columns = [standardize_header(col) for col in df.columns]
            has_key_info = False
            key_keywords = ['交易', '日期', '时间', '金额', '余额', '对方', '摘要', '备注', '流水']
            
            # 检查表头是否包含关键信息
            for col in df.columns:
                if any(keyword in col for keyword in key_keywords):
                    has_key_info = True
                    break
            
            # 如果表头没有关键信息，检查前50行数据
            if not has_key_info and len(df) > 0:
                sample_data = df.head(50).astype(str).values.flatten()
                for data in sample_data:
                    if any(keyword in data for keyword in key_keywords):
                        has_key_info = True
                        break
            
            # 如果没有关键信息，跳过处理
            if not has_key_info:
                self.logger.warning(f"【文本文件处理】{file_name} 不是有效的银行流水文件，缺少关键信息")
                return None
            
            # 处理文件数据
            return self._process_file_data(df, file_name)
        except Exception as e:
            self.logger.error(f"【文本文件处理失败】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_pdf_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        处理PDF文件，提取表格数据并进行标准化
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        
        try:
            self.logger.info(f"【PDF文件处理】开始处理：{file_name}")
            
            # 导入PDF处理模块
            from src.file_processing.pdf_processing import process_pdf_file
            
            # 处理PDF文件
            pdf_df = process_pdf_file(file_path, self.config)
            
            if pdf_df is None:
                self.logger.warning(f"【PDF文件处理】{file_name} 未提取到有效数据")
                return None
            
            # 处理文件数据
            processed_df = self._process_file_data(pdf_df, file_name)
            
            if processed_df is not None and not processed_df.empty:
                self.logger.info(f"【PDF文件处理】{file_name} 处理完成，共{len(processed_df)}行数据")
                return processed_df
            else:
                self.logger.warning(f"【PDF文件处理】{file_name} 处理后无有效数据")
                return None
                
        except Exception as e:
            self.logger.error(f"【PDF文件处理异常】{file_name}：{str(e)}", exc_info=True)
            return None

    def _process_excel_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        处理Excel文件，支持多sheet
        
        Args:
            file_path: 文件路径
            
        Returns:
            处理后的DataFrame或None
        """
        file_name = Path(file_path).name
        
        try:
            # 检测Excel文件是否加密
            if is_excel_encrypted(file_path):
                self.logger.warning(f"【Excel文件加密】{file_name}：文件已加密，无法处理")
                return None
            
            # 使用缓存避免重复读取同一个Excel文件
            if file_path not in self.excel_file_cache:
                self.excel_file_cache[file_path] = pd.ExcelFile(file_path)
            excel_file = self.excel_file_cache[file_path]
            # 更新访问时间
            self._update_cache_access("excel", file_path)
            
            all_sheets_data = []
            
            self.logger.debug(f"【Excel文件处理】{file_name} 包含 {len(excel_file.sheet_names)} 个sheet：{excel_file.sheet_names}")
            
            # 遍历所有sheet
            for sheet_name in excel_file.sheet_names:
                self.logger.info(f"【Excel文件处理】处理sheet：{sheet_name}")
                
                try:
                    # 读取单个sheet，保留合并单元格信息
                    # 首先获取sheet名称
                    df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    
                    # 如果数据为空，跳过处理
                    if df_raw.empty:
                        self.logger.warning(f"【Sheet有效性检查】{file_name} - {sheet_name} 为空，跳过处理")
                        continue
                    
                    # 首先在数据前50行中寻找最佳表头
                    best_header_row = 0
                    best_header_score = 0
                    best_header_headers = []
                    
                    # 最多检查前50行
                    max_check_rows = min(50, len(df_raw))
                    
                    self.logger.debug(f"【Excel表头识别】开始扫描前{max_check_rows}行寻找最佳表头")
                    
                    for i in range(max_check_rows):
                        # 尝试将第i行作为表头
                        header_candidate = df_raw.iloc[i]
                        
                        # 标准化表头候选
                        standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                        
                        # 计算表头得分
                        total_score = self._calculate_header_score(standardized_headers)
                        
                        # 更新最佳表头：只在得分更高时更新，得分相同时保留更前面的行
                        if total_score > best_header_score:
                            best_header_score = total_score
                            best_header_row = i
                            best_header_headers = standardized_headers
                    
                    # 输出最佳表头信息
                    self.logger.info(f"【表头分析】{file_name} - {sheet_name} 最佳表头行：第{best_header_row+1}行，得分：{best_header_score}")
                    self.logger.info(f"【表头分析】最佳表头：{best_header_headers}")
                    
                    # 使用最佳表头行
                    if best_header_row < len(df_raw):
                        df = df_raw.copy()
                        # 获取最佳表头
                        best_header = df.iloc[best_header_row].apply(standardize_header)
                        df.columns = best_header
                        df = df.iloc[best_header_row + 1:].reset_index(drop=True)
                        
                        # 处理文件数据，进行表头映射（临时映射，不进行关键列检查）
                        self.logger.info(f"【表头映射】{file_name} - {sheet_name} 开始临时表头映射")
                        
                        # 创建临时映射函数，跳过关键列检查
                        def temp_map(df_temp, file_id):
                            # 表头标准化
                            original_raw_headers = list(df_temp.columns)
                            df_temp.columns = [standardize_header(col) for col in df_temp.columns]
                            standardized_headers = list(df_temp.columns)
                            
                            # 优化重复列
                            df_temp = optimize_duplicate_columns(df_temp)
                            standardized_headers = list(df_temp.columns)
                            
                            # 映射到统一表头
                            mapped_df = pd.DataFrame(columns=self.unified_headers)
                            matched_cols = []
                            
                            # 收集所有可能的匹配
                            all_matches = []
                            
                            # 1. 收集所有精确匹配
                            for col in standardized_headers:
                                if col in self.header_mapping:
                                    unified_col = self.header_mapping[col]
                                    mapping_items = list(self.header_mapping.items())
                                    position = next((i for i, (k, v) in enumerate(mapping_items) if k == col), len(mapping_items))
                                    position_weight = max(0, 10 - position * 0.1)
                                    weight = 20 + position_weight
                                    
                                    all_matches.append({
                                        'original': col,
                                        'matched': unified_col,
                                        'method': '精确匹配',
                                        'weight': weight,
                                        'position': position
                                    })
                            
                            # 2. 收集所有关键词匹配
                            for col in standardized_headers:
                                if any(m['original'] == col and m['method'] == '精确匹配' for m in all_matches):
                                    continue
                                
                                for unified_col, keywords in self.keyword_mapping.items():
                                    for i, keyword in enumerate(keywords):
                                        if unified_col in self.exclusion_rules:
                                            exclusion_keywords = self.exclusion_rules[unified_col]
                                            if any(exclusion_keyword in col for exclusion_keyword in exclusion_keywords):
                                                continue
                                        
                                        score = 0
                                        if col == keyword:
                                            score = 10
                                        elif col.startswith(keyword) or col.endswith(keyword):
                                            score = 8
                                        elif keyword in col:
                                            keyword_chars = set(keyword)
                                            col_chars = set(col)
                                            if keyword_chars:
                                                common_chars = keyword_chars.intersection(col_chars)
                                                coverage_ratio = len(common_chars) / len(keyword_chars)
                                                if coverage_ratio > 0.5:
                                                    score = 5
                                        
                                        if score > 0:
                                            keyword_position_weight = max(0, 5 - i * 0.5)
                                            weight = score + keyword_position_weight
                                            
                                            all_matches.append({
                                                'original': col,
                                                'matched': unified_col,
                                                'method': '关键词匹配',
                                                'keyword': keyword,
                                                'weight': weight,
                                                'score': score,
                                                'keyword_position': i
                                            })
                            
                            # 3. 处理匹配冲突
                            matches_by_original = {}
                            for match in all_matches:
                                original_col = match['original']
                                if original_col not in matches_by_original:
                                    matches_by_original[original_col] = []
                                matches_by_original[original_col].append(match)
                            
                            selected_matches = []
                            for original_col, matches in matches_by_original.items():
                                sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                                best_match = sorted_matches[0]
                                selected_matches.append(best_match)
                            
                            # 4. 再次处理冲突
                            matches_by_unified = {}
                            for match in selected_matches:
                                unified_col = match['matched']
                                if unified_col not in matches_by_unified:
                                    matches_by_unified[unified_col] = []
                                matches_by_unified[unified_col].append(match)
                            
                            final_selected_matches = []
                            for unified_col, matches in matches_by_unified.items():
                                sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
                                best_match = sorted_matches[0]
                                final_selected_matches.append(best_match)
                            
                            selected_matches = final_selected_matches
                            
                            # 应用选定的匹配
                            for match in selected_matches:
                                original_col = match['original']
                                unified_col = match['matched']
                                if unified_col in mapped_df.columns:
                                    mapped_df[unified_col] = df_temp[original_col]
                                    matched_cols.append(match)
                            
                            # 排除全NA列
                            mapped_df = mapped_df.dropna(axis=1, how='all')
                            
                            # 将NaN替换为空字符串
                            mapped_df = mapped_df.fillna('').infer_objects()
                            
                            return mapped_df if not mapped_df.empty else None
                        
                        # 执行临时映射
                        temp_result = temp_map(df, f"{file_name} - {sheet_name} - 临时映射")
                        self.logger.info(f"【表头映射】{file_name} - {sheet_name} 临时表头映射完成，结果：{'成功' if temp_result is not None else '失败'}")
                        
                        # 检查表头映射结果
                        if temp_result is not None:
                            # 检查缺失的统一表头列
                            mapped_headers = list(temp_result.columns)
                            missing_headers = [col for col in self.unified_headers if col not in mapped_headers]
                            
                            self.logger.info(f"【表头映射分析】{file_name} - {sheet_name} 成功映射的统一表头列：{mapped_headers}")
                            self.logger.info(f"【表头映射分析】{file_name} - {sheet_name} 缺失的统一表头列：{missing_headers}")
                            
                            # 检查户主名称列是否在缺失列表中
                            has_missing_account_name = '户主名称' in missing_headers
                            self.logger.info(f"【表头映射分析】{file_name} - {sheet_name} 户主名称列是否缺失：{has_missing_account_name}")
                            
                            # 检查是否存在其他表头行评分大于等于最佳表头行评分
                            has_other_header_rows = False
                            if has_missing_account_name:
                                # 寻找最佳表头
                                current_best_header_row = 0
                                current_best_header_score = 0
                                
                                # 读取前50行寻找表头
                                header_df = pd.read_excel(
                                    excel_file,
                                    sheet_name=sheet_name,
                                    header=None,
                                    nrows=50
                                )
                                
                                for i in range(len(header_df)):
                                    header_candidate = header_df.iloc[i]
                                    standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                                    score = self._calculate_header_score(standardized_headers)
                                    
                                    if score > current_best_header_score:
                                        current_best_header_score = score
                                        current_best_header_row = i
                                
                                # 检查是否存在其他表头行评分大于等于最佳表头行评分
                                for i in range(len(header_df)):
                                    if i == current_best_header_row:
                                        continue
                                    header_candidate = header_df.iloc[i]
                                    standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                                    score = self._calculate_header_score(standardized_headers)
                                    if score >= current_best_header_score:
                                        has_other_header_rows = True
                                        break
                            
                            # 如果缺失的统一表头列中包含户主名称且存在其他表头行，尝试多账户处理
                            if has_missing_account_name and has_other_header_rows:
                                self.logger.info(f"【表头分析】{file_name} - {sheet_name} 表头映射后缺失户主名称列且存在其他表头行，尝试多账户处理")
                                self.logger.info(f"【多账户处理】开始多账户处理流程")
                                multi_account_df = self.multi_account_handler.process_multi_account_sheet(
                                    df_raw, 
                                    f"{file_name} - {sheet_name}"
                                )
                                
                                if multi_account_df is not None:
                                    # 多账户处理成功，直接处理结果
                                    self.logger.info(f"【多账户处理成功】{file_name} - {sheet_name} 采用多账户处理方式")
                                    sheet_result = self._process_file_data(multi_account_df, f"{file_name} - {sheet_name}")
                                    if sheet_result is not None and not sheet_result.empty:
                                        all_sheets_data.append(sheet_result)
                                        self.logger.info(f"【多账户处理】添加了{len(sheet_result)}行数据")
                                    continue
                                else:
                                    self.logger.info(f"【多账户处理】多账户处理失败，继续使用原处理流程")
                            else:
                                if has_missing_account_name:
                                    self.logger.info(f"【表头分析】{file_name} - {sheet_name} 表头映射后缺失户主名称列，但不存在其他表头行，使用原处理流程")
                                else:
                                    self.logger.info(f"【表头分析】{file_name} - {sheet_name} 表头映射后不缺失户主名称列，使用原处理流程")
                        else:
                            self.logger.warning(f"【表头映射】{file_name} - {sheet_name} 表头映射失败，继续使用原处理流程")
                    
                    # 处理合并单元格
                    # 对于Excel文件，pandas会自动填充合并单元格的值，但我们需要确保处理正确
                    try:
                        # 检查是否有合并单元格
                        workbook = excel_file.book
                        sheet = workbook[sheet_name]
                        
                        # 检测合并单元格（处理只读模式的情况）
                        has_merged_cells = False
                        merged_cells = []
                        
                        # 尝试不同的方式获取合并单元格
                        if hasattr(sheet, 'merged_cells'):
                            merged_cells = sheet.merged_cells
                            has_merged_cells = len(merged_cells) > 0
                        elif hasattr(sheet, '_merged_cells'):
                            merged_cells = sheet._merged_cells
                            has_merged_cells = len(merged_cells) > 0
                        elif hasattr(sheet, 'merged_cells_ranges'):
                            merged_cells = sheet.merged_cells_ranges
                            has_merged_cells = len(merged_cells) > 0
                        
                        if has_merged_cells:
                            merged_cells_list = list(merged_cells)
                            self.logger.debug(f"【Excel文件处理】{file_name} - {sheet_name} 检测到 {len(merged_cells_list)} 个合并单元格")
                            
                            # 验证合并单元格的数据完整性
                            # 遍历所有合并单元格
                            for merged_cell in merged_cells_list:
                                try:
                                    # 获取合并单元格的范围
                                    if hasattr(merged_cell, 'bounds'):
                                        min_row, min_col, max_row, max_col = merged_cell.bounds
                                    elif hasattr(merged_cell, 'min_row') and hasattr(merged_cell, 'max_row'):
                                        min_row = merged_cell.min_row
                                        min_col = merged_cell.min_col
                                        max_row = merged_cell.max_row
                                        max_col = merged_cell.max_col
                                    elif hasattr(merged_cell, 'start_row') and hasattr(merged_cell, 'end_row'):
                                        min_row = merged_cell.start_row + 1  # 调整为1-based索引
                                        min_col = merged_cell.start_column + 1
                                        max_row = merged_cell.end_row + 1
                                        max_col = merged_cell.end_column + 1
                                    else:
                                        # 尝试解析合并单元格范围字符串
                                        if hasattr(merged_cell, 'address'):
                                            # 对于xlrd的合并单元格
                                            addr = merged_cell.address
                                            # 解析范围字符串，如 "A1:B2"
                                            match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', addr)
                                            if match:
                                                min_col = self._excel_col_to_num(match.group(1))
                                                min_row = int(match.group(2))
                                                max_col = self._excel_col_to_num(match.group(3))
                                                max_row = int(match.group(4))
                                            else:
                                                continue
                                        else:
                                            continue
                                    
                                    # 检查合并单元格的值是否一致
                                    # 注意：openpyxl的行和列是从1开始的，而pandas是从0开始的
                                    # 因此需要转换为pandas的索引
                                    pandas_min_row = min_row - 1
                                    pandas_min_col = min_col - 1
                                    pandas_max_row = max_row - 1
                                    pandas_max_col = max_col - 1
                                    
                                    # 确保索引在有效范围内
                                    if pandas_min_row < 0 or pandas_max_row >= len(df_raw) or pandas_min_col < 0 or pandas_max_col >= len(df_raw.columns):
                                        self.logger.debug(f"【Excel合并单元格处理】{file_name} - {sheet_name} 合并单元格 [{min_row}:{max_row}, {min_col}:{max_col}] 超出数据范围，跳过处理")
                                        continue
                                    
                                    # 获取合并单元格的左上角值
                                    top_left_value = df_raw.iloc[pandas_min_row, pandas_min_col]
                                    
                                    # 检查合并范围内的所有单元格是否都等于左上角值
                                    # （pandas read_excel会自动填充合并单元格，但我们需要验证）
                                    merge_range = df_raw.iloc[pandas_min_row:pandas_max_row+1, pandas_min_col:pandas_max_col+1]
                                    if not (merge_range == top_left_value).all().all():
                                        self.logger.warning(f"【Excel合并单元格处理】{file_name} - {sheet_name} 合并单元格 [{min_row}:{max_row}, {min_col}:{max_col}] 值不一致，自动填充为左上角值")
                                        # 确保所有合并单元格的值一致
                                        df_raw.iloc[pandas_min_row:pandas_max_row+1, pandas_min_col:pandas_max_col+1] = top_left_value
                                        
                                    # 记录合并单元格信息，便于调试
                                    self.logger.debug(f"【Excel合并单元格处理】{file_name} - {sheet_name} 处理合并单元格 [{min_row}:{max_row}, {min_col}:{max_col}]，值：{top_left_value}")
                                except Exception as e:
                                    self.logger.debug(f"【Excel合并单元格处理】处理合并单元格异常：{str(e)}")
                                    continue
                    except Exception as e:
                        self.logger.debug(f"【Excel合并单元格处理】无法处理合并单元格：{str(e)}")
                        # 跳过合并单元格处理，继续执行后续逻辑
                    
                    # 现在在数据前50行中寻找最佳表头
                    best_header_row = 0
                    best_header_score = 0
                    
                    # 最多检查前50行
                    max_check_rows = min(50, len(df_raw))
                    
                    self.logger.debug(f"【Excel表头识别】开始扫描前{max_check_rows}行寻找最佳表头")
                    
                    for i in range(max_check_rows):
                        # 尝试将第i行作为表头
                        header_candidate = df_raw.iloc[i]
                        
                        # 标准化表头候选
                        standardized_headers = [standardize_header(str(h)) for h in header_candidate]
                        
                        # 计算表头得分
                        total_score = self._calculate_header_score(standardized_headers)
                        exact_match_score = sum(1 for h in standardized_headers if h in self.header_mapping)
                        keyword_match_score = 0
                        for h in standardized_headers:
                            for unified_col, keywords in self.keyword_mapping.items():
                                if any(keyword in h for keyword in keywords):
                                    keyword_match_score += 1
                                    break
                        
                        # 更新最佳表头：只在得分更高时更新，得分相同时保留更前面的行
                        if total_score > best_header_score:
                            best_header_score = total_score
                            best_header_row = i
                    
                    # 使用最佳表头行
                    if best_header_row < len(df_raw):
                        df = df_raw.copy()
                        # 获取最佳表头
                        best_header = df.iloc[best_header_row].apply(standardize_header)
                        df.columns = best_header
                        df = df.iloc[best_header_row + 1:].reset_index(drop=True)
                        
                        self.logger.info(f"【Excel表头识别完成】文件 {file_name} - {sheet_name} 最佳表头行：第{best_header_row+1}行")
                        self.logger.info(f"【Excel表头识别完成】最佳表头：{best_header.to_list()}")
                        self.logger.info(f"【Excel表头识别完成】表头得分：{best_header_score}")
                        print(f"【Excel表头识别完成】文件 {file_name} - {sheet_name} 最佳表头行：第{best_header_row+1}行")
                        print(f"【Excel表头识别完成】最佳表头：{best_header.to_list()}")
                    else:
                        # 如果没有找到合适的表头，返回None，放入待复核文件
                        self.logger.warning(f"【Excel处理】{file_name} - {sheet_name} 无法找到合适的表头，放入待复核文件")
                        continue
                    
                    # Sheet有效性检查
                    if df.empty:
                        self.logger.warning(f"【Sheet有效性检查】{file_name} - {sheet_name} 为空，跳过处理")
                        continue
                    
                    if len(df.columns) == 0:
                        self.logger.warning(f"【Sheet有效性检查】{file_name} - {sheet_name} 没有有效列，跳过处理")
                        continue
                    
                    # 检查是否有有效表头
                    if all(pd.isna(col) or col.strip() == '' for col in df.columns):
                        self.logger.warning(f"【Sheet有效性检查】{file_name} - {sheet_name} 没有有效表头，跳过处理")
                        continue
                    
                    self.logger.debug(f"【Excel文件处理】{file_name} - {sheet_name} 读取成功，共 {len(df)} 行，{len(df.columns)} 列")
                    
                    # 处理sheet数据
                    sheet_result = self._process_file_data(df, f"{file_name} - {sheet_name}")
                    if sheet_result is not None and not sheet_result.empty:
                        all_sheets_data.append(sheet_result)
                except Exception as e:
                    self.logger.error(f"【Sheet处理失败】{file_name} - {sheet_name}：{str(e)}")
                    continue
            
            # 合并所有sheet的数据
            if all_sheets_data:
                # 拼接前确保每个DataFrame都排除了全NA列
                cleaned_sheets_data = []
                for sheet_df in all_sheets_data:
                    if not sheet_df.empty:
                        # 排除全NA列，确保拼接时不会出现警告
                        cleaned_df = sheet_df.dropna(axis=1, how='all')
                        if not cleaned_df.empty:
                            cleaned_sheets_data.append(cleaned_df)
                
                if cleaned_sheets_data:
                    combined_df = pd.concat(cleaned_sheets_data, ignore_index=True)
                    self.logger.debug(f"【Excel文件处理】{file_name} 合并所有sheet数据，共 {len(combined_df)} 行")
                    return combined_df
                else:
                    return None
            else:
                return None
        except Exception as e:
            self.logger.error(f"【Excel文件处理失败】{file_name}：{str(e)}", exc_info=True)
            return None
    
    def _process_file_data(self, df: pd.DataFrame, file_identifier: str) -> Optional[pd.DataFrame]:
        """
        处理文件数据，包括表头标准化、映射到统一表头
        
        Args:
            df: 文件数据DataFrame
            file_identifier: 文件标识符（文件名或文件名+sheet名）
            
        Returns:
            处理后的DataFrame或None
        """
        # 表头标准化
        original_raw_headers = list(df.columns)
        df.columns = [standardize_header(col) for col in df.columns]
        standardized_headers = list(df.columns)
        
        # 记录原始表头和标准化表头
        self.logger.info(f"【表头识别】{file_identifier} 开始处理")
        self.logger.info(f"【表头识别】原始表头数量：{len(original_raw_headers)} 列")
        self.logger.info(f"【表头识别】原始表头：{original_raw_headers}")
        self.logger.debug(f"【表头标准化】{file_identifier} 标准化后的表头：{standardized_headers}")
        
        # 优化重复列
        df = optimize_duplicate_columns(df)
        standardized_headers = list(df.columns)
        self.logger.debug(f"【优化重复列】{file_identifier} 优化后共 {len(standardized_headers)} 列")
        
        # 映射到统一表头
        mapped_df = pd.DataFrame(columns=self.unified_headers)
        matched_cols = []
        exact_matches = []
        keyword_matches = []
        unmatched_cols = []
        
        # 收集所有可能的匹配
        all_matches = []
        
        self.logger.info(f"【表头映射】{file_identifier} 开始映射表头，共 {len(standardized_headers)} 列")
        print(f"【表头映射】{file_identifier} 开始映射表头，共 {len(standardized_headers)} 列")
        
        # 1. 收集所有精确匹配
        for col in standardized_headers:
            if col in self.header_mapping:
                unified_col = self.header_mapping[col]
                # 计算精确匹配权重：考虑词汇在映射库中的位置
                # 获取词汇在映射库中的索引，位置越靠前权重越高
                mapping_items = list(self.header_mapping.items())
                position = next((i for i, (k, v) in enumerate(mapping_items) if k == col), len(mapping_items))
                # 位置权重：位置越靠前，权重越高
                position_weight = max(0, 10 - position * 0.1)  # 最高10分，每靠后一位减0.1分
                # 精确匹配基础分 + 位置权重
                weight = 20 + position_weight  # 精确匹配基础分20分，确保高于关键词匹配
                
                all_matches.append({
                    'original': col,
                    'matched': unified_col,
                    'method': '精确匹配',
                    'weight': weight,
                    'position': position
                })
        
        # 2. 收集所有关键词匹配
        for col in standardized_headers:
            # 跳过已经有精确匹配的列
            if any(m['original'] == col and m['method'] == '精确匹配' for m in all_matches):
                continue
            
            # 计算每个统一列的匹配得分
            for unified_col, keywords in self.keyword_mapping.items():
                for i, keyword in enumerate(keywords):
                    # 检查排除规则
                    if unified_col in self.exclusion_rules:
                        exclusion_keywords = self.exclusion_rules[unified_col]
                        # 如果原始列包含任何排除关键词，则跳过该匹配
                        if any(exclusion_keyword in col for exclusion_keyword in exclusion_keywords):
                            continue
                    
                    # 计算匹配得分
                    score = 0
                    
                    # 1. 完全匹配，得10分
                    if col == keyword:
                        score = 10
                    # 2. 前缀/后缀匹配，得8分
                    elif col.startswith(keyword) or col.endswith(keyword):
                        score = 8
                    # 3. 包含匹配，得5分，但关键词的核心语义必须在原始字段中占比超过50%
                    elif keyword in col:
                        # 计算核心语义占比
                        # 使用基于字符的方法，计算关键词中字符在原始字段中的出现比例
                        # 对于中文，计算关键词中每个字符在原始字段中的出现次数
                        keyword_chars = set(keyword)
                        col_chars = set(col)
                        
                        # 计算关键词字符在原始字段中的覆盖率（核心语义占比）
                        if keyword_chars:
                            common_chars = keyword_chars.intersection(col_chars)
                            coverage_ratio = len(common_chars) / len(keyword_chars)
                            
                            # 只使用核心语义占比（字符覆盖率）判断，不考虑长度比例
                            if coverage_ratio > 0.5:
                                score = 5
                            else:
                                # 核心语义占比不足50%，不视为包含匹配
                                score = 0
                        else:
                            score = 0
                    
                    if score > 0:
                        # 计算关键词位置权重：位置越靠前权重越高
                        keyword_position_weight = max(0, 5 - i * 0.5)  # 最高5分，每靠后一位减0.5分
                        # 总权重 = 匹配得分 + 位置权重
                        weight = score + keyword_position_weight
                        
                        all_matches.append({
                            'original': col,
                            'matched': unified_col,
                            'method': '关键词匹配',
                            'keyword': keyword,
                            'weight': weight,
                            'score': score,
                            'keyword_position': i
                        })
        
        # 3. 处理匹配冲突：为每个原始列选择权重最高的匹配
        # 按原始列分组
        matches_by_original = {}
        for match in all_matches:
            original_col = match['original']
            if original_col not in matches_by_original:
                matches_by_original[original_col] = []
            matches_by_original[original_col].append(match)
        
        # 为每个原始列选择权重最高的匹配
        selected_matches = []
        for original_col, matches in matches_by_original.items():
            # 按权重降序排序
            sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
            # 选择权重最高的匹配
            best_match = sorted_matches[0]
            selected_matches.append(best_match)
        
        # 4. 再次处理冲突：确保每个统一列只被一个原始列匹配
        # 按统一列分组，选择权重最高的匹配
        matches_by_unified = {}
        for match in selected_matches:
            unified_col = match['matched']
            if unified_col not in matches_by_unified:
                matches_by_unified[unified_col] = []
            matches_by_unified[unified_col].append(match)
        
        # 为每个统一列选择权重最高的匹配
        final_selected_matches = []
        for unified_col, matches in matches_by_unified.items():
            # 按权重降序排序
            sorted_matches = sorted(matches, key=lambda x: x['weight'], reverse=True)
            # 选择权重最高的匹配
            best_match = sorted_matches[0]
            final_selected_matches.append(best_match)
        
        # 使用最终选择的匹配
        selected_matches = final_selected_matches
        
        # 4. 应用选定的匹配
        for match in selected_matches:
            original_col = match['original']
            unified_col = match['matched']
            method = match['method']
            
            if unified_col in mapped_df.columns:
                mapped_df[unified_col] = df[original_col]
                matched_cols.append(match)
                
                if method == '精确匹配':
                    exact_matches.append(f"{original_col} → {unified_col} (权重：{match['weight']:.2f})")
                    self.logger.info(f"【表头映射】精确匹配：{original_col} → {unified_col} (权重：{match['weight']:.2f})")
                    print(f"【表头映射】精确匹配：{original_col} → {unified_col} (权重：{match['weight']:.2f})")
                else:
                    keyword_matches.append(f"{original_col} → {unified_col} (关键词：{match['keyword']}, 权重：{match['weight']:.2f})")
                    self.logger.info(f"【表头映射】关键词匹配：{original_col} → {unified_col} (关键词：{match['keyword']}, 权重：{match['weight']:.2f})")
                    print(f"【表头映射】关键词匹配：{original_col} → {unified_col} (关键词：{match['keyword']}, 权重：{match['weight']:.2f})")
        
        # 5. 处理未匹配的列
        matched_original_cols = [m['original'] for m in selected_matches]
        for col in standardized_headers:
            if col not in matched_original_cols:
                unmatched_cols.append(col)
        
        # 记录未匹配的表头
        if unmatched_cols:
            self.logger.info(f"【表头映射】{file_identifier} 未匹配表头：{unmatched_cols}")
        
        # 记录表头映射结果汇总
        self.logger.info(f"【表头映射完成】{file_identifier} 映射结果汇总：")
        self.logger.info(f"【表头映射完成】精确匹配：{len(exact_matches)} 列")
        self.logger.info(f"【表头映射完成】关键词匹配：{len(keyword_matches)} 列")
        self.logger.info(f"【表头映射完成】未匹配：{len(unmatched_cols)} 列")
        self.logger.info(f"【表头映射完成】总计匹配：{len(matched_cols)} / {len(standardized_headers)} 列")
        
        # 记录详细匹配信息
        if exact_matches:
            self.logger.info(f"【表头映射详情】精确匹配列表：{exact_matches}")
        if keyword_matches:
            self.logger.info(f"【表头映射详情】关键词匹配列表：{keyword_matches}")
        
        print(f"【表头映射完成】{file_identifier} 映射结果汇总：")
        print(f"【表头映射完成】精确匹配：{len(exact_matches)} 列")
        print(f"【表头映射完成】关键词匹配：{len(keyword_matches)} 列")
        print(f"【表头映射完成】总计匹配：{len(matched_cols)} / {len(standardized_headers)} 列")
        
        # 记录成功映射的统一表头列
        mapped_headers = [m['matched'] for m in matched_cols]
        self.logger.info(f"【表头映射完成】成功映射的统一表头列：{mapped_headers}")
        self.logger.info(f"【表头映射完成】缺失的统一表头列：{[col for col in self.unified_headers if col not in mapped_headers]}")
        

        
        # 检查是否有有效数据
        if mapped_df.empty or mapped_df.isnull().all().all():
            self.logger.warning(f"【表头映射】{file_identifier} 无法匹配到有效表头，原始表头：{original_raw_headers}")
            # 表头识别失败，返回None，放入待复核文件
            return None
        
        # 初期关键列校验：检查表头映射结果是否包含关键列
        # 关键列校验功能置于表头映射环节之后
        
        # 收集成功映射的统一列
        mapped_unified_cols = [match['matched'] for match in matched_cols]
        
        # 检查关键列是否被实际映射
        has_time_col = '交易时间' in mapped_unified_cols
        has_date_col = '交易日期' in mapped_unified_cols
        has_amount_col = '交易金额' in mapped_unified_cols
        has_debit_col = '借方金额' in mapped_unified_cols
        has_credit_col = '贷方金额' in mapped_unified_cols
        
        # 记录关键列检测情况
        self.logger.info(f"【关键列检查】{file_identifier} 开始关键列检测")
        self.logger.info(f"【关键列检查】交易时间列：{'存在' if has_time_col else '不存在'}")
        self.logger.info(f"【关键列检查】交易日期列：{'存在' if has_date_col else '不存在'}")
        self.logger.info(f"【关键列检查】交易金额列：{'存在' if has_amount_col else '不存在'}")
        self.logger.info(f"【关键列检查】借方金额列：{'存在' if has_debit_col else '不存在'}")
        self.logger.info(f"【关键列检查】贷方金额列：{'存在' if has_credit_col else '不存在'}")
        self.logger.info(f"【关键列检查】成功映射的统一列：{mapped_unified_cols}")
        
        # 检查是否有可映射匹配到的关键列
        has_any_key_col = has_time_col or has_date_col or has_amount_col or has_debit_col or has_credit_col
        
        # 若无一可映射匹配，则该表（sheet）视为无效数据
        if not has_any_key_col:
            self.logger.warning(f"【关键列检查】{file_identifier} 没有可映射匹配到的关键列，视为无效数据")
            return None
        
        # 设定条件
        # 条件（1）：能映射到交易日期或交易时间
        condition1 = has_time_col or has_date_col
        # 条件（2）：能同时映射到借方金额和贷方金额
        condition2 = has_debit_col and has_credit_col
        # 条件（3）：能映射到交易金额
        condition3 = has_amount_col
        
        # 记录条件检查结果
        self.logger.info(f"【关键列检查】条件（1）- 能映射到交易日期或交易时间：{'满足' if condition1 else '不满足'}")
        self.logger.info(f"【关键列检查】条件（2）- 能同时映射到借方金额和贷方金额：{'满足' if condition2 else '不满足'}")
        self.logger.info(f"【关键列检查】条件（3）- 能映射到交易金额：{'满足' if condition3 else '不满足'}")
        
        # 仅当同时满足条件（1）（2）、同时满足条件（1）（3）、同时满足条件（1）（2）（3）时，判定该表（sheet）数据为有效数据
        is_valid = (condition1 and condition2) or (condition1 and condition3)
        
        # 记录最终判定结果
        if is_valid:
            self.logger.info(f"【关键列检查】{file_identifier} 关键列检测通过，视为有效数据")
            # 记录满足的条件组合
            if condition1 and condition2 and condition3:
                self.logger.info(f"【关键列检查】满足条件组合：（1）（2）（3）")
            elif condition1 and condition2:
                self.logger.info(f"【关键列检查】满足条件组合：（1）（2）")
            elif condition1 and condition3:
                self.logger.info(f"【关键列检查】满足条件组合：（1）（3）")
        else:
            # 否则该表（sheet）视为无效数据
            missing_conditions = []
            if not condition1:
                missing_conditions.append("能映射到交易日期或交易时间")
            if not (condition2 or condition3):
                missing_conditions.append("能同时映射到借方金额和贷方金额，或能映射到交易金额")
            self.logger.warning(f"【关键列检查】{file_identifier} 缺少必要条件：{', '.join(missing_conditions)}，视为无效数据")
            return None
        
        # 移除数据标准化步骤，直接使用原始数据
        
        # 检查交易时间列数据是否为空，若为空则尝试使用其他日期列作为备用
        if has_time_col:
            # 增强空值处理策略：当交易时间列为空时，尝试使用其他日期列作为备用
            if '交易时间' in mapped_df.columns:
                # 检查是否有其他日期列
                date_columns = ['记账日期', '入账日期', '发生日期']
                for date_col in date_columns:
                    if date_col in mapped_df.columns:
                        # 尝试使用其他日期列填充交易时间列的空值
                        mask = mapped_df['交易时间'].isnull() | (mapped_df['交易时间'] == '')
                        if mask.any():
                            self.logger.info(f"【关键列检查】{file_identifier} 交易时间列存在空值，使用 {date_col} 列作为备用")
                            mapped_df.loc[mask, '交易时间'] = mapped_df.loc[mask, date_col]
                            break
        
        self.logger.info(f"【关键列检查】{file_identifier} 关键列检查通过，共保留 {len(mapped_df)} 行有效数据")
        
        # 6. 将特定列转换为字符串类型，避免科学计数法
        text_columns = ['本方账号', '本方卡号', '交易对方账号', '交易对方卡号', '交易流水号', '柜员号', 'IP地址', 'MAC地址']
        for col in text_columns:
            if col in mapped_df.columns:
                mapped_df[col] = mapped_df[col].astype(str)
        
        # 7. 排除全NA列，避免后续拼接警告
        mapped_df = mapped_df.dropna(axis=1, how='all')
        
        # 将NaN替换为空字符串，保持原始数据为空的地方为空
        mapped_df = mapped_df.fillna('').infer_objects()
        
        return mapped_df if not mapped_df.empty else None
