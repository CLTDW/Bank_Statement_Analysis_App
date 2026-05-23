import os
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Union
import logging
import multiprocessing
import time

# 尝试导入psutil，如果不可用则使用默认值
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    psutil = None
    PSUTIL_AVAILABLE = False

from src.utils import collect_all_supported_files
from src.logging import setup_logger
from src.core.config_manager import ConfigManager
from src.core.mapping_manager import MappingManager
from src.core.file_processor import FileProcessor

class ResourceMonitor:
    """
    资源监控器，用于实时监控CPU和内存使用率
    """
    def __init__(self, logger, max_cpu_usage=90, max_memory_usage=90):
        """
        初始化资源监控器
        
        Args:
            logger: 日志记录器
            max_cpu_usage: 最大CPU使用率（%）
            max_memory_usage: 最大内存使用率（%）
        """
        self.logger = logger
        self.max_cpu_usage = max_cpu_usage
        self.max_memory_usage = max_memory_usage
        self.is_psutil_available = PSUTIL_AVAILABLE
    
    def check_resources(self) -> Dict[str, bool]:
        """
        检查当前系统资源使用情况
        
        Returns:
            资源使用情况字典，包含CPU和内存是否超限
        """
        if not self.is_psutil_available:
            return {
                "cpu_ok": True,
                "memory_ok": True,
                "cpu_usage": 0,
                "memory_usage": 0
            }
        
        try:
            # 获取CPU使用率
            cpu_usage = psutil.cpu_percent(interval=0.1)
            
            # 获取内存使用率
            memory = psutil.virtual_memory()
            memory_usage = memory.percent
            
            # 检查是否超限
            cpu_ok = cpu_usage <= self.max_cpu_usage
            memory_ok = memory_usage <= self.max_memory_usage
            
            return {
                "cpu_ok": cpu_ok,
                "memory_ok": memory_ok,
                "cpu_usage": cpu_usage,
                "memory_usage": memory_usage
            }
        except Exception as e:
            self.logger.warning(f"【资源监控】获取资源使用情况失败：{str(e)}")
            return {
                "cpu_ok": True,
                "memory_ok": True,
                "cpu_usage": 0,
                "memory_usage": 0
            }
    
    def wait_for_resources(self, timeout=60):
        """
        等待资源使用情况恢复正常
        
        Args:
            timeout: 超时时间（秒）
        
        Returns:
            是否在超时前恢复正常
        """
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            status = self.check_resources()
            
            if status["cpu_ok"] and status["memory_ok"]:

                return True
            
            # 资源紧张，等待一段时间
            wait_time = min(1, (70 - max(status["cpu_usage"], status["memory_usage"])) / 10)
            wait_time = max(0.1, wait_time)
            
            self.logger.warning(f"【资源监控】资源使用紧张 - CPU: {status['cpu_usage']:.1f}%, 内存: {status['memory_usage']:.1f}%, 等待 {wait_time:.1f}秒")
            time.sleep(wait_time)
        
        self.logger.error("【资源监控】等待资源超时，继续执行")
        return False
    
    def log_resource_usage(self):
        """
        记录当前资源使用情况
        """
        status = self.check_resources()


class BankStatementAggregator:
    def __init__(self, config: Dict):
        self.config = config
        # 初始化配置管理器
        self.config_manager = ConfigManager()
        self.config_manager.config = config
        
        # 初始化日志记录器
        self.logger = setup_logger(config)
        
        self.logger.info(f"初始化银行流水汇总程序，配置文件：{config.get('config_file', '默认配置')}")
        self.logger.info(f"输入目录：{self.config_manager.get_base_dir()}")
        self.logger.info(f"输出文件：{self.config_manager.get_output_file()}")
        self.logger.info(f"模板文件：{self.config_manager.get_template_file()}")

        # 检测电脑配置
        self.system_config = self._detect_system_config()
        
        # 初始化资源监控器，设置最大资源使用阈值为90%
        self.resource_monitor = ResourceMonitor(self.logger, max_cpu_usage=90, max_memory_usage=90)
        
        # 生成最优资源分配方案
        self.resource_plan = self._generate_resource_plan()

        # 初始化映射管理器
        self.mapping_manager = MappingManager(config)
        
        # 初始化用户输入的对象名称
        self.object_names = []
        
        # 初始化文件处理器
        self.file_processor = FileProcessor(config, self.mapping_manager, self.object_names)
        
        # 核心属性
        self.expected_columns = self.mapping_manager.expected_columns
        
        # 初始化处理报告
        self.process_report: Dict[str, Union[int, Dict, List]] = {
            "success": 0,
            "fail": 0,
            "total": 0,
            "total_rows": 0,
            "failed_files": [],
            "review_files": []  # 待复核文件列表
        }


        
        # 创建待复核文件文件夹
        self.review_folder = "待复核文件"
        self.review_folder_path = os.path.join(self.config_manager.get_base_dir(), self.review_folder)
        os.makedirs(self.review_folder_path, exist_ok=True)

    def _validate_base_dir(self) -> None:
        """
        校验基础目录
        """
        base_dir = self.config_manager.get_base_dir()
        if not base_dir.exists():
            self.logger.error(f"基础目录不存在：{base_dir}")
            raise FileNotFoundError(f"基础目录不存在：{base_dir}")
        if not base_dir.is_dir():
            self.logger.error(f"{base_dir}不是有效目录")
            raise NotADirectoryError(f"{base_dir}不是有效目录")
        self.logger.info(f"基础目录校验通过：{base_dir}")

    def _validate_file_exists(self, file_path: Path, file_desc: str) -> None:
        """
        校验文件是否存在
        """
        if not file_path.exists():
            self.logger.error(f"{file_desc}不存在：{file_path}")
            raise FileNotFoundError(f"{file_desc}不存在：{file_path}")
        self.logger.info(f"{file_desc}校验通过：{file_path}")

    def _backup_original_files(self, files: List[str]):
        """
        备份所有原始文件到带有时间戳的目录
        """
        import shutil
        from datetime import datetime
        
        # 创建带有时间戳的备份目录
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        backup_dir = os.path.join(self.config_manager.get_base_dir(), "原始文件备份", timestamp)
        os.makedirs(backup_dir, exist_ok=True)
        
        self.logger.info(f"【自动备份】将备份所有原始文件到：{backup_dir}")
        print(f"【自动备份】将备份所有原始文件到：{backup_dir}")
        
        # 复制所有文件到备份目录
        for file_path in files:
            try:
                # 保持相对路径结构
                relative_path = os.path.relpath(file_path, self.config_manager.get_base_dir())
                backup_file_path = os.path.join(backup_dir, relative_path)
                
                # 创建父目录
                os.makedirs(os.path.dirname(backup_file_path), exist_ok=True)
                
                # 复制文件
                shutil.copy2(file_path, backup_file_path)
                self.logger.debug(f"【备份成功】{file_path} → {backup_file_path}")
            except Exception as e:
                self.logger.error(f"【备份失败】{file_path}：{str(e)}")
                print(f"【备份失败】{file_path}：{str(e)}")
    
    def _move_to_review_folder(self, file_path: str, reason: str):
        """
        将文件移动到待复核文件夹
        """
        import shutil
        from datetime import datetime
        
        # 确保待复核文件夹存在
        review_folder_path = os.path.join(self.config_manager.get_base_dir(), self.review_folder)
        os.makedirs(review_folder_path, exist_ok=True)
        
        # 生成带有时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_name = os.path.basename(file_path)
        name, ext = os.path.splitext(file_name)
        review_file_name = f"{name}_{timestamp}{ext}"
        review_file_path = os.path.join(review_folder_path, review_file_name)
        
        try:
            # 复制文件到待复核目录
            shutil.copy2(file_path, review_file_path)
            # 记录待复核文件
            self.process_report["review_files"].append({
                "file": str(file_path),
                "reason": reason,
                "review_file": review_file_name
            })
            self.logger.info(f"【待复核】已将 {file_name} 移动到待复核文件夹，原因：{reason}")
            print(f"【待复核】已将 {file_name} 移动到待复核文件夹，原因：{reason}")
        except Exception as e:
            self.logger.error(f"【待复核失败】{file_name}：{str(e)}")
            print(f"【待复核失败】{file_name}：{str(e)}")
    
    def _detect_system_config(self) -> Dict:
        """
        检测电脑系统配置
        
        Returns:
            系统配置字典，包含CPU、内存、磁盘等信息
        """
        self.logger.info("【系统配置检测】开始检测电脑硬件配置")
        
        config = {
            "cpu": {
                "logical_cores": multiprocessing.cpu_count()
            },
            "memory": {
                "total_gb": 0,
                "available_gb": 0
            },
            "disk": {
                "read_speed_mb_s": 0
            },
            "is_low_config": False
        }
        
        if PSUTIL_AVAILABLE:
            try:
                # 获取内存信息
                memory = psutil.virtual_memory()
                config["memory"]["total_gb"] = memory.total / (1024 ** 3)
                config["memory"]["available_gb"] = memory.available / (1024 ** 3)
                
                # 简单评估磁盘速度
                disk_speed = self._estimate_disk_speed()
                config["disk"]["read_speed_mb_s"] = disk_speed
                
                # 判断是否为低配置电脑
                config["is_low_config"] = (
                    config["cpu"]["logical_cores"] <= 4 and
                    config["memory"]["total_gb"] <= 8 and
                    config["memory"]["available_gb"] <= 2
                )
                
            except Exception as e:
                self.logger.warning(f"【系统配置检测】获取系统信息失败：{str(e)}")
        else:
            self.logger.warning("【系统配置检测】psutil不可用，无法获取完整系统信息")
        

        
        return config
    
    def _estimate_disk_speed(self) -> float:
        """
        简单评估磁盘读取速度
        
        Returns:
            预估的磁盘读取速度（MB/s）
        """
        try:
            # 创建一个临时文件并读取，评估速度
            import tempfile
            import os
            
            # 创建一个10MB的临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False)
            temp_file_path = temp_file.name
            temp_file.close()
            
            # 写入10MB数据
            with open(temp_file_path, 'wb') as f:
                f.write(b'0' * (10 * 1024 * 1024))
            
            # 读取文件并计时
            start_time = time.time()
            with open(temp_file_path, 'rb') as f:
                _ = f.read()
            end_time = time.time()
            
            # 计算速度
            elapsed_time = end_time - start_time
            if elapsed_time > 0:
                speed = (10) / elapsed_time  # 10MB / 时间
            else:
                speed = 0
            
            # 清理临时文件
            os.unlink(temp_file_path)
            
            return speed
        except Exception:
            return 0
    
    def _generate_resource_plan(self) -> Dict:
        """
        基于系统配置生成最优资源分配方案
        
        Returns:
            资源分配方案字典
        """
        self.logger.info("【资源方案生成】开始生成最优资源分配方案")
        
        plan = {
            "parallel": {
                "excel_threads": 0,
                "batch_size": 0
            },
            "memory": {
                "chunk_size": 0,
                "cache_size": 0
            },
            "disk": {
                "max_concurrent_io": 0
            }
        }
        
        # 获取系统配置
        cpu_cores = self.system_config["cpu"]["logical_cores"]
        total_memory = self.system_config["memory"]["total_gb"]
        available_memory = self.system_config["memory"]["available_gb"]
        is_low_config = self.system_config["is_low_config"]
        
        # 基于CPU核心数计算线程/进程数
        if is_low_config:
            # 低配置电脑：保守设置
            plan["parallel"]["excel_threads"] = min(4, cpu_cores)
            plan["parallel"]["batch_size"] = 50
            plan["memory"]["chunk_size"] = 5000
            plan["memory"]["cache_size"] = 50
            plan["disk"]["max_concurrent_io"] = 2
        else:
            # 高配置电脑：更积极的设置
            # Excel/CSV/TXT：IO密集型，使用更多线程
            plan["parallel"]["excel_threads"] = min(32, max(4, cpu_cores * 2))
            plan["parallel"]["batch_size"] = 100
            plan["memory"]["chunk_size"] = 10000
            plan["memory"]["cache_size"] = 100
            plan["disk"]["max_concurrent_io"] = 4
        
        # 动态资源分配：根据当前系统资源使用情况调整
        # 首先检查当前资源使用情况
        current_status = self.resource_monitor.check_resources()
        current_cpu_usage = current_status["cpu_usage"]
        current_memory_usage = current_status["memory_usage"]
        
        # 计算可用资源百分比
        available_cpu_percent = 90 - current_cpu_usage  # 最大90%
        available_memory_percent = 90 - current_memory_usage  # 最大90%
        
        # 基于可用资源计算线程数
        max_cpu_usage = min(available_cpu_percent, 90)
        plan["parallel"]["excel_threads"] = min(
            plan["parallel"]["excel_threads"],
            int(cpu_cores * max_cpu_usage / 100) * 2  # IO密集型可以使用更多线程
        )
        
        # 基于可用内存计算批处理大小
        max_memory_usage = min(available_memory_percent, 90)
        available_memory_mb = available_memory * 1024
        memory_limit_mb = available_memory_mb * max_memory_usage / 100
        
        # 调整批处理大小以适应内存限制
        # 假设每个文件处理需要50MB内存
        estimated_memory_per_file = 50
        max_files_in_batch = int(memory_limit_mb / estimated_memory_per_file)
        plan["parallel"]["batch_size"] = min(plan["parallel"]["batch_size"], max_files_in_batch, 200)
        
        # 确保批处理大小至少为10
        plan["parallel"]["batch_size"] = max(10, plan["parallel"]["batch_size"])
        

        
        return plan
    
    def _calculate_optimal_workers(self, file_count: int, file_type: str = "excel") -> int:
        """
        根据系统资源和文件类型计算最佳并行度
        
        Args:
            file_count: 文件数量
            file_type: 文件类型，可选值："excel"（Excel/CSV/TXT）
            
        Returns:
            最佳工作线程/进程数
        """
        try:
            # 检查当前资源使用情况
            resource_status = self.resource_monitor.check_resources()
            
            # 计算当前可用资源百分比
            available_cpu_percent = 90 - resource_status["cpu_usage"]  # 最大90%
            available_memory_percent = 90 - resource_status["memory_usage"]  # 最大90%
            
            # 基于可用资源动态调整并行度
            if available_cpu_percent < 10 or available_memory_percent < 10:
                # 资源非常紧张，大幅减少并行度
                self.logger.warning(f"【并行处理】资源严重紧张，大幅降低并行度 - CPU: {resource_status['cpu_usage']:.1f}%, 内存: {resource_status['memory_usage']:.1f}%")
                return max(1, self.resource_plan["parallel"]["excel_threads"] // 4)
            elif available_cpu_percent < 30 or available_memory_percent < 30:
                # 资源紧张，适当减少并行度
                self.logger.warning(f"【并行处理】资源紧张，调整并行度 - CPU: {resource_status['cpu_usage']:.1f}%, 内存: {resource_status['memory_usage']:.1f}%")
                return max(1, self.resource_plan["parallel"]["excel_threads"] // 2)
            else:
                # 资源充足，使用计划的并行度
                optimal_workers = self.resource_plan["parallel"]["excel_threads"]
                
                # 根据文件数量调整
                optimal_workers = min(file_count, optimal_workers)
                
                # 确保至少有1个工作线程
                optimal_workers = max(1, optimal_workers)
                
                self.logger.info(f"【并行处理】计算最佳{file_type}处理线程数：{optimal_workers}")
                return optimal_workers
        except Exception as e:
            self.logger.warning(f"【并行处理】计算最佳工作线程数异常：{str(e)}，使用默认值")
            return max(1, multiprocessing.cpu_count() // 2)

    def run(self):
        """
        运行银行流水汇总程序
        """
        self.logger.info("\n" + "=" * 80)
        self.logger.info("【程序启动】开始运行银行流水汇总程序")
        print("\n" + "=" * 80)
        print("【程序启动】开始运行银行流水汇总程序")
        
        # 初始化校验
        self.logger.info("【初始化校验】开始校验基础配置")
        print("【初始化校验】开始校验基础配置")
        
        self._validate_base_dir()
        self._validate_file_exists(self.config_manager.get_template_file(), "流水汇总模板")
        
        self.logger.info("【初始化校验】基础配置校验完成")
        print("【初始化校验】基础配置校验完成")
        
        # 获取用户输入的对象名称
        self.logger.info("【用户输入】开始获取此次汇总流水的对象名称")
        print("\n" + "=" * 60)
        print("【用户输入】请输入此次汇总流水的所有对象名称")
        print("【提示】多个名称请用逗号、顿号作为分隔符，例如：张三、李四,王五")
        user_input = input("请输入对象名称：").strip()
        
        # 解析用户输入的对象名称
        self.object_names = []
        if user_input:
            # 按逗号和顿号分割
            import re
            object_names_list = re.split(r'[,，、]', user_input)
            # 去除空字符串和空格
            self.object_names = [name.strip() for name in object_names_list if name.strip()]
            self.logger.info(f"【用户输入】获取到对象名称：{self.object_names}")
            print(f"【用户输入】获取到对象名称：{self.object_names}")
        else:
            self.logger.info("【用户输入】未输入对象名称")
            print("【用户输入】未输入对象名称")
            self.object_names = []

        # 重新初始化文件处理器，使用用户输入的对象名称
        self.file_processor = FileProcessor(self.config, self.mapping_manager, self.object_names)
        self.logger.info("【文件处理器】已使用用户输入的对象名称重新初始化")

        # 安全提示
        self.logger.warning("⚠️  重要提示：程序将自动备份所有原始银行流水文件！")
        print("⚠️  重要提示：程序将自动备份所有原始银行流水文件！")

        # 收集所有支持的文件
        self.logger.info("【文件收集】开始收集所有支持的银行流水文件")
        print("【文件收集】开始收集所有支持的银行流水文件")
        
        files = collect_all_supported_files(self.config_manager.get_base_dir(), self.config)
        
        self.logger.info(f"【文件收集完成】共找到 {len(files)} 个文件需要处理")
        self.logger.debug(f"【文件列表】{files}")
        print(f"【文件收集完成】共找到 {len(files)} 个文件需要处理")
        
        # 自动备份原始文件
        if files:
            self.logger.info(f"【文件备份】开始备份 {len(files)} 个原始文件")
            print(f"【文件备份】开始备份 {len(files)} 个原始文件")
            self._backup_original_files(files)
            self.logger.info("【文件备份】原始文件备份完成")
            print("【文件备份】原始文件备份完成")
        else:
            self.logger.warning("【文件收集】未找到任何需要处理的文件")
            print("【文件收集】未找到任何需要处理的文件")
            return

        # 收集所有处理结果
        all_results = []
        
        # 检查是否启用并行处理
        use_parallel = self.config.get('parallel', {}).get('enabled', False)
        
        # 所有文件都作为Excel/CSV/TXT文件处理
        excel_files = files
        
        self.logger.info(f"【文件分类】Excel/CSV/TXT文件：{len(excel_files)} 个")
        print(f"【文件分类】Excel/CSV/TXT文件：{len(excel_files)} 个")

        # 处理Excel/CSV/TXT文件
        if excel_files:
            self.logger.info(f"\n{'=' * 60}")
            self.logger.info(f"【Excel文件处理】开始处理 {len(excel_files)} 个Excel/CSV/TXT文件")
            print(f"\n{'=' * 60}")
            print(f"【Excel文件处理】开始处理 {len(excel_files)} 个Excel/CSV/TXT文件")
            
            # 分批处理
            batch_size = self.resource_plan["parallel"]["batch_size"]
            for i in range(0, len(excel_files), batch_size):
                batch_files = excel_files[i:i+batch_size]
                self.logger.info(f"【批次处理】处理第 {i//batch_size + 1} 批，共 {len(batch_files)} 个文件")
                
                # 检查资源状态
                resource_status = self.resource_monitor.check_resources()
                if not resource_status["cpu_ok"] or not resource_status["memory_ok"]:
                    self.logger.warning(f"【资源监控】资源紧张，等待资源释放 - CPU: {resource_status['cpu_usage']:.1f}%, 内存: {resource_status['memory_usage']:.1f}%")
                    self.resource_monitor.wait_for_resources()
                
                # 计算最佳并行度
                if use_parallel:
                    # 重新生成资源方案，确保基于最新的系统资源状态
                    self.resource_plan = self._generate_resource_plan()
                    max_workers = self._calculate_optimal_workers(len(batch_files), "excel")
                else:
                    max_workers = 1
                
                self.logger.info(f"【处理配置】并行处理：{use_parallel}，最大工作线程：{max_workers}")
                
                if use_parallel:
                    # 使用线程池处理
                    from tqdm import tqdm
                    from concurrent.futures import ThreadPoolExecutor, as_completed
                    
                    # 改进任务分配策略，按文件大小排序，先处理小文件
                    def get_file_size(file_path):
                        try:
                            return os.path.getsize(file_path)
                        except:
                            return 0
                    
                    # 按文件大小排序，小文件优先
                    sorted_batch_files = sorted(batch_files, key=get_file_size)
                    
                    def process_file_wrapper(file_path):
                        return self.process_single_file(file_path)
                    
                    # 改进线程池管理，使用上下文管理器确保资源释放
                    with ThreadPoolExecutor(max_workers=max_workers) as executor:
                        # 提交任务并记录开始时间
                        task_start_times = {}
                        futures = {}
                        
                        for file_path in sorted_batch_files:
                            future = executor.submit(process_file_wrapper, file_path)
                            futures[future] = file_path
                            task_start_times[future] = time.time()
                        
                        # 监控任务执行情况
                        completed_count = 0
                        total_tasks = len(futures)
                        
                        # 收集所有任务结果
                        task_results = []
                        
                        for future in tqdm(as_completed(futures), total=total_tasks, desc="【Excel文件处理进度】"):
                            file_path = futures[future]
                            completed_count += 1
                            
                            # 计算任务执行时间
                            execution_time = time.time() - task_start_times[future]
                            
                            try:
                                df = future.result()
                                task_results.append({
                                    'file_path': file_path,
                                    'df': df,
                                    'execution_time': execution_time,
                                    'error': None
                                })
                            except Exception as e:
                                task_results.append({
                                    'file_path': file_path,
                                    'df': None,
                                    'execution_time': execution_time,
                                    'error': str(e)
                                })
                        
                        # 按文件顺序输出日志
                        self.logger.info(f"【并行处理完成】批次处理完成，共{completed_count}个文件")
                        
                        # 输出处理结果
                        success_count = 0
                        fail_count = 0
                        
                        for result in task_results:
                            file_path = result['file_path']
                            execution_time = result['execution_time']
                            
                            if result['error']:
                                # 处理异常
                                self.logger.error(f"【处理异常】{file_path}：{result['error']}，耗时{execution_time:.2f}秒")
                                print(f"【处理异常】{file_path}：{result['error']}")
                                self.process_report["fail"] += 1
                                self.process_report["failed_files"].append(f"{file_path}：{result['error']}")
                                fail_count += 1
                            else:
                                df = result['df']
                                if df is not None and len(df) > 0:
                                    # 处理成功
                                    all_results.append(df)
                                    self.logger.info(f"【处理成功】{file_path}：处理完成，共{len(df)}行数据，耗时{execution_time:.2f}秒")
                                    success_count += 1
                                else:
                                    # 处理失败
                                    ext = os.path.splitext(file_path)[1].lower()
                                    if ext == '.pdf':
                                        # 检查失败文件列表，确定失败原因
                                        file_name = os.path.basename(file_path)
                                        # 查找失败原因
                                        fail_reason = "PDF表格提取失败，无有效数据"
                                        for failed_file in self.process_report["failed_files"]:
                                            if file_name in failed_file and "表格提取成功但无有效数据" in failed_file:
                                                fail_reason = "PDF表格提取成功但无有效数据"
                                                break
                                        self.logger.warning(f"【处理失败】{file_path}：{fail_reason}，移至待复核，耗时{execution_time:.2f}秒")
                                        self._move_to_review_folder(file_path, fail_reason)
                                    else:
                                        self.logger.warning(f"【处理失败】{file_path}：处理结果为空，移至待复核，耗时{execution_time:.2f}秒")
                                        self._move_to_review_folder(file_path, "处理结果为空")
                                    fail_count += 1
                        
                        total_time = time.time() - task_start_times[list(futures.keys())[0]]
                        self.logger.info(f"【批次处理统计】成功：{success_count}个，失败：{fail_count}个，总耗时：{total_time:.2f}秒")
                else:
                    # 串行处理
                    from tqdm import tqdm
                    
                    for file_path in tqdm(batch_files, desc="【Excel文件处理进度】"):
                        df = self.process_single_file(file_path)
                        if df is not None and len(df) > 0:
                            all_results.append(df)
                        else:
                            # 检查是否为PDF文件，提供更具体的原因
                            ext = os.path.splitext(file_path)[1].lower()
                            if ext == '.pdf':
                                # 检查失败文件列表，确定失败原因
                                file_name = os.path.basename(file_path)
                                # 查找失败原因
                                fail_reason = "PDF表格提取失败，无有效数据"
                                for failed_file in self.process_report["failed_files"]:
                                    if file_name in failed_file and "表格提取成功但无有效数据" in failed_file:
                                        fail_reason = "PDF表格提取成功但无有效数据"
                                        break
                                self.logger.warning(f"【处理失败】{file_path}：{fail_reason}，移至待复核")
                                self._move_to_review_folder(file_path, fail_reason)
                            else:
                                self.logger.warning(f"【处理失败】{file_path}：处理结果为空，移至待复核")
                                self._move_to_review_folder(file_path, "处理结果为空")
        
        # 处理PDF/HTML文件

        
        self.logger.info(f"【文件处理完成】共处理 {len(files)} 个文件，成功 {self.process_report['success']} 个，失败 {self.process_report['fail']} 个")
        print(f"【文件处理完成】共处理 {len(files)} 个文件，成功 {self.process_report['success']} 个，失败 {self.process_report['fail']} 个")

        # 保存所有结果
        if all_results:
            # 检查资源状态
            resource_status = self.resource_monitor.check_resources()
            if not resource_status["cpu_ok"] or not resource_status["memory_ok"]:
                self.logger.warning(f"【资源监控】资源紧张，等待资源释放 - CPU: {resource_status['cpu_usage']:.1f}%, 内存: {resource_status['memory_usage']:.1f}%")
                self.resource_monitor.wait_for_resources()
            
            self.logger.info(f"\n{'=' * 80}")
            self.logger.info(f"【结果保存】开始保存所有处理结果")
            print(f"\n{'=' * 80}")
            print(f"【结果保存】开始保存所有处理结果")
            
            # 预处理：确保每个DataFrame只包含预期的列，并按预期顺序排列，同时保留数据来源列
            self.logger.info(f"【数据预处理】开始预处理 {len(all_results)} 个处理结果")
            print(f"【数据预处理】开始预处理 {len(all_results)} 个处理结果")
            
            processed_results = []
            total_rows = 0
            
            for i, df in enumerate(all_results):
                self.logger.info(f"【数据预处理】处理第 {i+1}/{len(all_results)} 个结果，共 {len(df)} 行数据")
                
                # 直接使用原始DataFrame，不基于预期列创建新的DataFrame
                # 只保留原始文件中存在的数据列
                formatted_df = df.copy()
                
                # 确保非空
                if not formatted_df.empty:
                    # 排除全NA列，避免concat警告，但保留数据来源列
                    if '数据来源' in formatted_df.columns:
                        # 只排除非数据来源的全NA列
                        non_source_cols = [col for col in formatted_df.columns if col != '数据来源']
                        # 先获取非数据来源列的非全NA列
                        non_source_valid_cols = formatted_df[non_source_cols].dropna(axis=1, how='all').columns.tolist()
                        # 重新构建DataFrame，保留数据来源列和有效的非数据来源列
                        formatted_df = formatted_df[['数据来源'] + non_source_valid_cols]
                    else:
                        # 如果没有数据来源列，只排除全NA列
                        formatted_df = formatted_df.dropna(axis=1, how='all')
                    
                    processed_results.append(formatted_df)
                    total_rows += len(formatted_df)
                    
                    self.logger.info(f"【数据预处理】第 {i+1} 个结果预处理完成，共 {len(formatted_df)} 行，{len(formatted_df.columns)} 列")
                    self.logger.debug(f"【预处理后列】{list(formatted_df.columns)}")
            
            self.logger.info(f"【数据预处理完成】共预处理 {len(processed_results)} 个结果，总计 {total_rows} 行数据")
            print(f"【数据预处理完成】共预处理 {len(processed_results)} 个结果，总计 {total_rows} 行数据")
            
            # 合并所有结果
            if processed_results:
                self.logger.info(f"【数据合并】开始合并 {len(processed_results)} 个处理结果")
                print(f"【数据合并】开始合并 {len(processed_results)} 个处理结果")
                
                # 合并所有DataFrame，使用sort=False避免FutureWarning
                final_df = pd.concat(processed_results, ignore_index=True, sort=False)
                
                # 将NaN替换为空字符串，保持原始数据为空的地方为空
                final_df = final_df.fillna('')
                
                self.logger.info(f"【数据合并完成】合并后的数据：{len(final_df)}行，{len(final_df.columns)}列")
                self.logger.debug(f"【合并后列】{list(final_df.columns)}")
                print(f"【数据合并完成】合并后的数据：{len(final_df)}行，{len(final_df.columns)}列")
                
                # 创建一个空的DataFrame，使用预期的列
                self.logger.info(f"【数据映射】开始将合并后的数据映射到预期列")
                expected_df = pd.DataFrame(columns=self.expected_columns)
                
                # 将合并后的DataFrame中的列映射到预期列，同时保留数据来源列
                mapped_cols = []
                for col in final_df.columns:
                    if col in expected_df.columns:
                        expected_df[col] = final_df[col]
                        mapped_cols.append(col)
                    elif col == '数据来源':
                        # 保留数据来源列
                        expected_df[col] = final_df[col]
                        mapped_cols.append(col)
                
                self.logger.info(f"【数据映射完成】成功映射 {len(mapped_cols)} 列数据")
                print(f"【数据映射完成】成功映射 {len(mapped_cols)} 列数据")
                
                # 添加序号列
                self.logger.info("【数据处理】添加序号列")
                expected_df['序号'] = range(1, len(expected_df) + 1)
                
                # 不再过滤核心列为空的数据，保留所有行
                self.logger.info("【行级过滤】保留所有行数据，不再过滤核心列为空的数据")
                filtered_df = expected_df.copy()
                
                # 只检查是否有核心列，不进行行级过滤
                has_time_col = '交易时间' in expected_df.columns
                has_date_col = '交易日期' in expected_df.columns
                has_amount_col = '交易金额' in expected_df.columns
                has_debit_col = '借方金额' in expected_df.columns
                has_credit_col = '贷方金额' in expected_df.columns
                
                # 检查是否有时间相关列
                has_time_related_col = has_time_col or has_date_col
                # 检查是否有金额相关列
                has_amount_related_col = has_amount_col or (has_debit_col and has_credit_col)
                
                if not has_time_related_col or not has_amount_related_col:
                    # 如果缺少核心列，记录警告但仍然保留数据
                    self.logger.warning("【行级过滤】汇总数据缺少核心列，但仍然保留所有数据")
                    print("【行级过滤】汇总数据缺少核心列，但仍然保留所有数据")
                
                self.logger.info(f"【行级过滤】保留所有 {len(filtered_df)} 行数据")
                print(f"【行级过滤】保留所有 {len(filtered_df)} 行数据")
                
                # 确保数据来源列出现在表格末尾
                if '数据来源' in filtered_df.columns:
                    # 将数据来源列移到最后
                    cols = [col for col in filtered_df.columns if col != '数据来源']
                    cols.append('数据来源')
                    filtered_df = filtered_df[cols]
                    self.logger.info("【数据来源】将数据来源列移到表格末尾")
                    print("【数据来源】将数据来源列移到表格末尾")
                
                # 保存结果
                output_file = self.config_manager.get_output_file()
                # 添加时间戳到输出文件名
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d%H%M")
                # 兼容Python 3.7的方式修改文件名
                new_stem = f"{output_file.stem}{timestamp}"
                output_file = output_file.parent / f"{new_stem}{output_file.suffix}"
                
                # 同时在TEST文件夹中保存一份
                test_dir = Path('TEST')
                test_output_file = test_dir / f"{new_stem}{output_file.suffix}"
                
                self.logger.info(f"【结果保存】开始保存处理结果到 {output_file}")
                self.logger.info(f"【结果保存】同时保存到TEST文件夹：{test_output_file}")
                print(f"【结果保存】开始保存处理结果到 {output_file}")
                print(f"【结果保存】同时保存到TEST文件夹：{test_output_file}")
                
                # 检查数据量，实现分批处理
                total_rows = len(filtered_df)
                # 根据总行数决定分批策略
                if total_rows < 300000:
                    batch_size = total_rows  # 低于30万行，全部存为一个文件
                else:
                    batch_size = 200000  # 高于30万行，每20万行存为一个文件
                
                # 定义需要设置为文本格式的列
                text_columns = ['本方账号', '本方卡号', '交易对方账号', '交易对方卡号', '交易流水号', '柜员号', 'IP地址', 'MAC地址']
                
                if total_rows > 300000:
                    # 数据量较大，分批保存
                    batch_size = 200000  # 每20万行存为一个文件
                    self.logger.info(f"【分批处理】数据量较大（{total_rows}行），开始分批保存，每批{batch_size}行")
                    print(f"【分批处理】数据量较大（{total_rows}行），开始分批保存，每批{batch_size}行")
                    
                    # 计算批次数
                    batch_count = (total_rows + batch_size - 1) // batch_size
                    
                    for i in range(batch_count):
                        start_row = i * batch_size
                        end_row = min((i + 1) * batch_size, total_rows)
                        batch_df = filtered_df.iloc[start_row:end_row]
                        
                        # 生成批次文件名
                        batch_file = output_file.with_stem(f"{output_file.stem}_batch{i+1}")
                        test_batch_file = test_output_file.with_stem(f"{test_output_file.stem}_batch{i+1}")
                        
                        self.logger.info(f"【分批处理】保存第{i+1}/{batch_count}批，行数：{start_row+1}-{end_row}")
                        print(f"【分批处理】保存第{i+1}/{batch_count}批，行数：{start_row+1}-{end_row}")
                        
                        try:
                            # 尝试保存为Excel，设置文本格式
                            from openpyxl import Workbook
                            from openpyxl.utils import get_column_letter
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            from openpyxl.styles import numbers
                            
                            # 创建工作簿
                            wb = Workbook()
                            ws = wb.active
                            
                            # 写入表头
                            headers = list(batch_df.columns)
                            ws.append(headers)
                            
                            # 写入数据
                            for r in dataframe_to_rows(batch_df, index=False, header=False):
                                ws.append(r)
                            
                            # 设置文本格式
                            text_col_indices = []
                            for col_idx, col in enumerate(headers):
                                if col in text_columns:
                                    text_col_indices.append(col_idx)
                            
                            for col_idx in text_col_indices:
                                col_letter = get_column_letter(col_idx + 1)
                                for row in range(2, ws.max_row + 1):
                                    cell = ws[f"{col_letter}{row}"]
                                    cell.number_format = '@'  # 设置为文本格式
                            
                            # 保存文件
                            wb.save(batch_file)
                            self.logger.info(f"【分批处理】第{i+1}批保存成功：{batch_file}")
                            print(f"【分批处理】第{i+1}批保存成功：{batch_file}")
                            
                            # 同时保存到TEST文件夹
                            wb.save(test_batch_file)
                            self.logger.info(f"【分批处理】第{i+1}批保存到TEST文件夹成功：{test_batch_file}")
                            print(f"【分批处理】第{i+1}批保存到TEST文件夹成功：{test_batch_file}")
                        except Exception as e:
                            # 保存失败时尝试保存为CSV
                            csv_file = batch_file.with_suffix('.csv')
                            test_csv_file = test_batch_file.with_suffix('.csv')
                            self.logger.warning(f"【分批处理】Excel保存失败，尝试保存为CSV：{str(e)}")
                            print(f"【分批处理】Excel保存失败，尝试保存为CSV：{str(e)}")
                            batch_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                            self.logger.info(f"【分批处理】第{i+1}批保存为CSV成功：{csv_file}")
                            print(f"【分批处理】第{i+1}批保存为CSV成功：{csv_file}")
                            
                            # 同时保存到TEST文件夹
                            batch_df.to_csv(test_csv_file, index=False, encoding='utf-8-sig')
                            self.logger.info(f"【分批处理】第{i+1}批保存到TEST文件夹为CSV成功：{test_csv_file}")
                            print(f"【分批处理】第{i+1}批保存到TEST文件夹为CSV成功：{test_csv_file}")
                    
                    self.logger.info(f"【分批处理完成】共保存 {total_rows} 行数据到 {batch_count} 个文件")
                    print(f"【分批处理完成】共保存 {total_rows} 行数据到 {batch_count} 个文件")
                else:
                    # 数据量较小，直接保存为一个文件
                    try:
                        # 尝试保存为Excel，设置文本格式
                        from openpyxl import Workbook
                        from openpyxl.utils import get_column_letter
                        from openpyxl.utils.dataframe import dataframe_to_rows
                        from openpyxl.styles import numbers
                        
                        # 创建工作簿
                        wb = Workbook()
                        ws = wb.active
                        
                        # 写入表头
                        headers = list(filtered_df.columns)
                        ws.append(headers)
                        
                        # 写入数据
                        for r in dataframe_to_rows(filtered_df, index=False, header=False):
                            ws.append(r)
                        
                        # 设置文本格式
                        text_col_indices = []
                        for col_idx, col in enumerate(headers):
                            if col in text_columns:
                                text_col_indices.append(col_idx)
                        
                        for col_idx in text_col_indices:
                            col_letter = get_column_letter(col_idx + 1)
                            for row in range(2, ws.max_row + 1):
                                cell = ws[f"{col_letter}{row}"]
                                cell.number_format = '@'  # 设置为文本格式
                        
                        # 保存文件
                        wb.save(output_file)
                        self.logger.info(f"【结果保存完成】共保存 {total_rows} 行数据到 {output_file}")
                        self.logger.info(f"【输出格式】包含 {len(filtered_df.columns)} 列，列顺序与模板一致")
                        print(f"【结果保存完成】共保存 {total_rows} 行数据到 {output_file}")
                        print(f"【输出格式】包含 {len(filtered_df.columns)} 列，列顺序与模板一致")
                        
                        # 同时保存到TEST文件夹
                        wb.save(test_output_file)
                        self.logger.info(f"【结果保存完成】共保存 {total_rows} 行数据到TEST文件夹：{test_output_file}")
                        print(f"【结果保存完成】共保存 {total_rows} 行数据到TEST文件夹：{test_output_file}")
                    except Exception as e:
                        # 保存失败时尝试保存为CSV
                        csv_file = output_file.with_suffix('.csv')
                        test_csv_file = test_output_file.with_suffix('.csv')
                        self.logger.warning(f"【结果保存】Excel保存失败，尝试保存为CSV：{str(e)}")
                        print(f"【结果保存】Excel保存失败，尝试保存为CSV：{str(e)}")
                        filtered_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
                        self.logger.info(f"【结果保存完成】共保存 {total_rows} 行数据到 {csv_file}")
                        print(f"【结果保存完成】共保存 {total_rows} 行数据到 {csv_file}")
                        
                        # 同时保存到TEST文件夹
                        filtered_df.to_csv(test_csv_file, index=False, encoding='utf-8-sig')
                        self.logger.info(f"【结果保存完成】共保存 {total_rows} 行数据到TEST文件夹：{test_csv_file}")
                        print(f"【结果保存完成】共保存 {total_rows} 行数据到TEST文件夹：{test_csv_file}")
            else:
                self.logger.warning(f"【结果保存】没有有效的数据可保存")
                print(f"【结果保存】没有有效的数据可保存")
        else:
            self.logger.warning(f"【结果保存】没有有效的数据可保存")
            print(f"【结果保存】没有有效的数据可保存")

        # 生成报告
        self.logger.info("【报告生成】开始生成处理报告")
        print("【报告生成】开始生成处理报告")
        
        self._generate_report()
        
        self.logger.info("【报告生成完成】处理报告已生成")
        print("【报告生成完成】处理报告已生成")
        
        # 记录最终资源使用情况
        self.resource_monitor.log_resource_usage()
        
        self.logger.info(f"\n{'=' * 80}")
        self.logger.info("【程序结束】银行流水汇总处理完成！")
        self.logger.info(f"【处理统计】总计：{self.process_report['total']} 个文件，成功：{self.process_report['success']} 个，失败：{self.process_report['fail']} 个")
        self.logger.info(f"【处理统计】处理总行数：{self.process_report['total_rows']} 行")
        
        if self.process_report['failed_files']:
            self.logger.info(f"【处理统计】失败文件数：{len(self.process_report['failed_files'])} 个")
        
        if self.process_report['review_files']:
            self.logger.info(f"【处理统计】待复核文件数：{len(self.process_report['review_files'])} 个")
        
        print(f"\n{'=' * 80}")
        print("【程序结束】银行流水汇总处理完成！")
        print(f"【处理统计】总计：{self.process_report['total']} 个文件，成功：{self.process_report['success']} 个，失败：{self.process_report['fail']} 个")
        print(f"【处理统计】处理总行数：{self.process_report['total_rows']} 行")
        
        if self.process_report['failed_files']:
            print(f"【处理统计】失败文件数：{len(self.process_report['failed_files'])} 个")
        
        if self.process_report['review_files']:
            print(f"【处理统计】待复核文件数：{len(self.process_report['review_files'])} 个")

    def process_single_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        处理单个文件（支持多sheet/CSV/TXT，原有逻辑不变）
        """
        file_name = os.path.basename(file_path)
        self.logger.info(f"【开始处理】{file_name}")
        print(f"【开始处理】{file_name}")
        self.process_report["total"] += 1

        # 初始化最终结果
        final_df = None

        try:
            # 使用文件处理器处理文件
            df = self.file_processor.process_file(file_path)
            if df is not None:
                final_df = df

            if final_df is not None and len(final_df) > 0:
                # 添加数据来源列，将文件完整路径汇入其中
                final_df['数据来源'] = file_path
                self.logger.info(f"【数据来源】{file_name}：添加数据来源列，路径：{file_path}")
                
                self.process_report["success"] += 1
                self.process_report["total_rows"] += len(final_df)
                self.logger.info(f"【处理成功】{file_name}：处理完成，共{len(final_df)}行数据")
                print(f"【处理成功】{file_name}：处理完成，共{len(final_df)}行数据")
                return final_df
            else:
                self.process_report["fail"] += 1
                # 检查是否为PDF文件，提供更具体的失败原因
                ext = os.path.splitext(file_path)[1].lower()
                if ext == '.pdf':
                    # 表格提取成功但无有效数据
                    fail_reason = f"{file_name}：表格提取成功但无有效数据"
                else:
                    # 其他文件处理失败
                    fail_reason = f"{file_name}：处理失败，无有效数据"
                # 只添加一个失败原因条目
                self.process_report["failed_files"].append(fail_reason)
                self.logger.warning(f"【处理失败】{file_name}：无有效数据")
                print(f"【处理失败】{file_name}：无有效数据")
                # 不再在这里移动文件，由run方法统一处理
                return None

        except Exception as e:
            self.process_report["fail"] += 1
            self.process_report["failed_files"].append(f"{file_name}：{str(e)}")
            self.logger.error(f"【处理异常】{file_name}：{str(e)}", exc_info=True)
            print(f"【处理异常】{file_name}：{str(e)}")
            # 不再在这里移动文件，由run方法统一处理
            return None



    def _generate_report(self):
        """
        生成处理报告
        """
        report_file = self.config_manager.get_report_file()
        # 同时在TEST文件夹中生成报告
        test_dir = Path('TEST')
        test_report_file = test_dir / report_file.name
        
        # 生成报告内容
        report_content = []
        report_content.append("银行流水汇总处理报告\n")
        report_content.append("=" * 50 + "\n")
        report_content.append(f"总文件数：{self.process_report['total']}\n")
        report_content.append(f"成功处理：{self.process_report['success']}\n")
        report_content.append(f"处理失败：{self.process_report['fail']}\n")
        report_content.append(f"处理总行数：{self.process_report['total_rows']}\n")
        report_content.append("\n失败文件列表：\n")
        for failed_file in self.process_report['failed_files']:
            report_content.append(f"  - {failed_file}\n")
        report_content.append("\n待复核文件列表：\n")
        for review_file in self.process_report['review_files']:
            report_content.append(f"  - {review_file['file']}：{review_file['reason']} → {review_file['review_file']}\n")
        
        # 写入报告文件
        report_text = "".join(report_content)
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_text)
        
        # 同时写入TEST文件夹
        with open(test_report_file, 'w', encoding='utf-8') as f:
            f.write(report_text)
        
        self.logger.info(f"【报告生成完成】处理报告已保存到：{report_file}")
        self.logger.info(f"【报告生成完成】处理报告已保存到TEST文件夹：{test_report_file}")
        print(f"【报告生成完成】处理报告已保存到：{report_file}")
        print(f"【报告生成完成】处理报告已保存到TEST文件夹：{test_report_file}")
